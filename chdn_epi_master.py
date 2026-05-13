"""
CHDN EPI Complete Data Transformation Pipeline
Consolidates all transformation steps into one master script
"""

from __future__ import annotations

import argparse
import re
import sys
import warnings
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

SCRIPT_DIR = Path(__file__).resolve().parent
PRIMARY_DATA_FILE_CANDIDATES = [
    SCRIPT_DIR / "CHDN_EPI_clean.xlsx",
    SCRIPT_DIR / "CHDN_clean" / "CHDN_EPI_clean.xlsx",
]
DATA_FILE = PRIMARY_DATA_FILE_CANDIDATES[0]
OUTPUT_FILE = SCRIPT_DIR / "CHDN dataset_long.xlsx"
FALLBACK_DATA_FILE = SCRIPT_DIR / "CHDN dataset.xlsm"
SECOND_FALLBACK_DATA_FILE = SCRIPT_DIR / "CHDN_dataset_copy.xlsm"


def _sheet_exists(workbook_path: Path, sheet_name: str) -> bool:
    """Return True if workbook exists and contains the requested sheet."""
    if not workbook_path.exists():
        return False
    try:
        wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            return sheet_name in wb.sheetnames
        finally:
            wb.close()
    except Exception:
        return False


def get_sheet_source_file(sheet_name: str) -> Path:
    """Resolve which source workbook should be used for a specific sheet."""
    candidates = [DATA_FILE, *PRIMARY_DATA_FILE_CANDIDATES, FALLBACK_DATA_FILE, SECOND_FALLBACK_DATA_FILE]
    for path in candidates:
        if _sheet_exists(path, sheet_name):
            if path != DATA_FILE:
                print(f"Using fallback source for '{sheet_name}': {path.name}")
            return path
    return DATA_FILE


def get_data_file() -> Path:
    """Get the data file, using fallback if primary is locked."""
    for primary_path in PRIMARY_DATA_FILE_CANDIDATES:
        if not primary_path.exists():
            continue
        try:
            pd.read_excel(primary_path, sheet_name="EPI-Child", nrows=1)
            if primary_path != PRIMARY_DATA_FILE_CANDIDATES[0]:
                print(f"Using alternate primary data file: {primary_path.name}")
            return primary_path
        except Exception:
            continue
    
    if FALLBACK_DATA_FILE.exists():
        print(f"Using fallback data file: {FALLBACK_DATA_FILE.name}")
        return FALLBACK_DATA_FILE

    if SECOND_FALLBACK_DATA_FILE.exists():
        print(f"Using fallback data file: {SECOND_FALLBACK_DATA_FILE.name}")
        return SECOND_FALLBACK_DATA_FILE
    
    return DATA_FILE


def print_header(title: str) -> None:
    print("\n" + "=" * 60)
    print(title)
    print("=" * 60)


def normalize_text(value) -> str:
    """Normalize text by stripping extra spaces."""
    if pd.isna(value):
        return ""
    return " ".join(str(value).split())


def get_period(date) -> str | None:
    """Convert date to period format (Q1_2025, etc.) using custom quarter cutoffs:
    Q1: Dec 21 (prev year) – Mar 20
    Q2: Mar 21 – Jun 20
    Q3: Jun 21 – Sep 20
    Q4: Sep 21 – Dec 20
    """
    if pd.isna(date):
        return None
    d = pd.to_datetime(date)
    y = d.year

    q1_start = pd.Timestamp(year=y - 1, month=12, day=21)
    q1_end = pd.Timestamp(year=y, month=3, day=20)
    if q1_start <= d <= q1_end:
        return f"Q1_{y}"

    q2_start = pd.Timestamp(year=y, month=3, day=21)
    q2_end = pd.Timestamp(year=y, month=6, day=20)
    if q2_start <= d <= q2_end:
        return f"Q2_{y}"

    q3_start = pd.Timestamp(year=y, month=6, day=21)
    q3_end = pd.Timestamp(year=y, month=9, day=20)
    if q3_start <= d <= q3_end:
        return f"Q3_{y}"

    q4_start = pd.Timestamp(year=y, month=9, day=21)
    q4_end = pd.Timestamp(year=y, month=12, day=20)
    if q4_start <= d <= q4_end:
        return f"Q4_{y}"

    return f"Q1_{y + 1}"


def get_age_group(age) -> str | None:
    """Categorize age into groups: U1 (0-11mo), U5 (12-59mo), >5 (60+mo)."""
    if pd.isna(age):
        return None
    try:
        age_val = int(float(age))
    except Exception:
        return None
    if age_val <= 11:
        return "U1"
    elif age_val <= 59:
        return "U5"
    else:
        return ">5"


def initialize_output_file() -> None:
    """Create or clear output file."""
    if OUTPUT_FILE.exists():
        OUTPUT_FILE.unlink()
    wb = openpyxl.Workbook()
    # Keep the default sheet (required by openpyxl) - it will be overwritten later
    wb.active.title = "Temp"
    wb.save(OUTPUT_FILE)
    print(f"Created output file: {OUTPUT_FILE}")


def step_transform_epi_long() -> None:
    """Transform EPI-Child from wide to long format."""
    print_header("Step 1: Transform EPI-Child Data to Long Format")

    dose_mappings = {
        "BCG": {
            "dose": "BCG dose",
            "dose_other": "BCG other (Y/N)",
            "reporting_month": "BCG reporting month",
            "age": "BCG Age ",
            "source": "BCG Source",
        },
        "OPV1": {
            "dose": "OPV first time dose",
            "dose_other": "OPV first time dose other",
            "reporting_month": "OPV first time dose reporting month",
            "age": "OPV1 Age ",
            "source": "OPV1 Source",
        },
        "OPV2": {
            "dose": "OPV second time dose",
            "dose_other": "OPV second time dose other",
            "reporting_month": "OPV second time dose reporting month",
            "age": "OPV2 Age",
            "source": "OPV2 Source",
        },
        "OPV3": {
            "dose": "OPV third time dose",
            "dose_other": "OPV third time dose other",
            "reporting_month": "OPV third time dose reporting month",
            "age": "OPV3 Age",
            "source": "OPV3 Source",
        },
        "Penta1": {
            "dose": "Penta first time dose",
            "dose_other": "Penta first time dose other",
            "reporting_month": "Penta first time dose reporting month",
            "age": "Penta1 Age",
            "source": "Penta1 Source",
        },
        "Penta2": {
            "dose": "Penta second time dose",
            "dose_other": "Penta second time dose other",
            "reporting_month": "Penta second time dose reporting month",
            "age": "Penta2 Age",
            "source": "Penta2 Source",
        },
        "Penta3": {
            "dose": "Penta third time dose",
            "dose_other": "Penta third time dose other",
            "reporting_month": "Penta third time dose reporting month",
            "age": "Penta3 Age",
            "source": "Penta3 Source",
        },
        "MMR1": {
            "dose": "MMR first timeadose",
            "dose_other": "MMR first time dose other",
            "reporting_month": "MMR first time dose reporting month",
            "age": "MMR1 Age",
            "source": "MMR1 Source",
        },
        "MMR2": {
            "dose": "MMR second time dose",
            "dose_other": "MMR second time dose other",
            "reporting_month": "MMR second time dose reporting month",
            "age": "MMR2 Age",
            "source": "MMR2 source",
        },
        "JE1": {
            "dose": "JE first time dose",
            "dose_other": "JE first time dose other",
            "reporting_month": "JE first time dose reporting month",
            "age": "JE1 Age",
            "source": "JE1 Source",
        },
        "JE2": {
            "dose": "JE second time dose",
            "dose_other": "JE second time dose other",
            "reporting_month": "JE second time dose reporting month",
            "age": "JE2 Age",
            "source": "JE2 Source",
        },
        "IPV": {
            "dose": "IPV dose",
            "dose_other": "IPV dose other",
            "reporting_month": "IPV_reporting_month",
            "age": "IPV Age",
            "source": "IPV Source",
        },
        "Rota1": {
            "dose": "Rota first time dose",
            "dose_other": "Rota first time dose other",
            "reporting_month": "Rota first time dose reporting month",
            "age": "Rota1 Age",
            "source": "Rota1 Source",
        },
        "Rota2": {
            "dose": "Rota second time dose",
            "dose_other": "Rota second time dose other",
            "reporting_month": "Rota second time dose reporting month",
            "age": "Rota2 Age ",
            "source": "Rota2 Source",
        },
    }

    keep_columns = [
        "No",
        "registered_date",
        "township_name_MIMU",
        "Team Name",
        "Session Point",
        "children_name",
        "address",
        "children_code(T/C Code)",
        "Gender",
        "IDP (Y/N)",
        "Date of Birth",
        "comments",
        "Age at first visit",
    ]

    time_columns = ["2024 Q4"]
    drop_extraneous_time_cols = [
        "2025 Q1", "2025 Q2", "2025 Q3", "2025 Q4", "2025 S1", "2025 S2",
        "At least one dose In 2025", "CompleteInQ4 2024",
        "2026 Q1", "2026 Q2", "2026 Q3", "2026 Q4", "2026 S1", "2026 S2",
        "At least one dose In 2026",
    ]
    completion_cols_2025 = [
        "CompleteInQ1 2025", "CompleteInQ2 2025", "CompleteInQ3 2025", "CompleteInQ4 2025"
    ]
    fallback_quarterly_2025 = ["2025 Q1", "2025 Q2", "2025 Q3", "2025 Q4"]
    completion_cols_2026 = [
        "CompleteInQ1 2026", "CompleteInQ2 2026", "CompleteInQ3 2026", "CompleteInQ4 2026",
        "CompleteIn 2026", "CompleteInS1_2026", "CompleteInS2_2026"
    ]

    print("Reading Excel file...")
    df = pd.read_excel(DATA_FILE, sheet_name="EPI-Child")
    print(f"Original data shape: {df.shape}")
    print(f"Original columns: {len(df.columns)}")

    if "vaccine_dose" in df.columns:
        print("Detected existing long-form sheet; copying to new file.")
        ordered_cols = [c for c in keep_columns if c in df.columns] + \
            [c for c in time_columns if c in df.columns] + \
            ["vaccine_dose", "age_at_dose", "reporting_month", "source"]
        if "completion_status" in df.columns and "completion_status" not in ordered_cols:
            ordered_cols.append("completion_status")
        long_df = df[ordered_cols]
    else:
        print("Building long-form from wide sheet...")
        long_frames = []

        for vaccine_name, cols_map in dose_mappings.items():
            required_cols = [
                cols_map["dose"],
                cols_map["dose_other"],
                cols_map["age"],
                cols_map["reporting_month"],
                cols_map["source"],
            ]
            missing = [c for c in required_cols if c not in df.columns]
            if missing:
                print(f"Warning: missing columns for {vaccine_name}: {missing}")
                continue

            subset_cols = [
                c for c in keep_columns + time_columns + completion_cols_2025 + fallback_quarterly_2025 + completion_cols_2026 + required_cols
                if c in df.columns
            ]
            subdf = df[subset_cols].copy()
            subdf.rename(
                columns={
                    cols_map["dose"]: "dose",
                    cols_map["dose_other"]: "dose_other",
                    cols_map["age"]: "age_at_dose",
                    cols_map["reporting_month"]: "reporting_month",
                    cols_map["source"]: "source",
                },
                inplace=True,
            )
            subdf["vaccine_dose"] = vaccine_name

            ordered_cols = [c for c in keep_columns if c in subdf.columns] + \
                [c for c in time_columns if c in subdf.columns] + \
                [c for c in completion_cols_2025 + fallback_quarterly_2025 + completion_cols_2026 if c in subdf.columns] + \
                ["vaccine_dose", "age_at_dose", "reporting_month", "source"]
            subdf = subdf[ordered_cols]
            long_frames.append(subdf)
            print(f"  {vaccine_name}: {len(subdf)} rows")

        if not long_frames:
            raise SystemExit("No vaccine frames created; check column mappings.")

        long_df = pd.concat(long_frames, ignore_index=True)

    if "dose" in long_df.columns:
        long_df = long_df.drop(columns=["dose"])
    if "dose_other" in long_df.columns:
        long_df = long_df.drop(columns=["dose_other"])

    print("\nFiltering out rows with age_at_dose=1111 and source='Not Received Yet'...")
    rows_before_filter = len(long_df)
    long_df = long_df[~((long_df["age_at_dose"] == 1111) & (long_df["source"] == "Not Received Yet"))]
    print(f"Removed {rows_before_filter - len(long_df)} rows")

    print("Filtering out rows with blank children_name...")
    rows_before_name_filter = len(long_df)
    long_df = long_df[long_df["children_name"].notna()]
    long_df = long_df[long_df["children_name"].astype(str).str.strip() != ""]
    print(f"Removed {rows_before_name_filter - len(long_df)} rows due to blank children_name")

    print("\nCombining completion columns for 2025 and 2026...")
    combination_priority = completion_cols_2025 + fallback_quarterly_2025 + completion_cols_2026
    existing_completion_cols = [col for col in combination_priority if col in long_df.columns]
    print(f"Found completion columns: {existing_completion_cols}")

    def get_completion_status(row):
        for col in existing_completion_cols:
            val = row[col]
            if pd.isna(val) or val == "" or val == 0:
                continue

            val_str = str(val).strip()
            core_val = val_str.split("complete in")[0].strip() if "complete in" in val_str.lower() else val_str

            if "Q1 2025" in col:
                return f"{core_val} complete in Q1_2025"
            if "Q2 2025" in col:
                return f"{core_val} complete in Q2_2025"
            if "Q3 2025" in col:
                return f"{core_val} complete in Q3_2025"
            if "Q4 2025" in col:
                return f"{core_val} complete in Q4_2025"
            if "Q1 2026" in col:
                return f"{core_val} complete in Q1_2026"
            if "Q2 2026" in col:
                return f"{core_val} complete in Q2_2026"
            if "Q3 2026" in col:
                return f"{core_val} complete in Q3_2026"
            if "Q4 2026" in col:
                return f"{core_val} complete in Q4_2026"
            if "S1_2026" in col:
                return f"{core_val} complete in S1_2026"
            if "S2_2026" in col:
                return f"{core_val} complete in S2_2026"
            if "CompleteIn 2026" == col:
                return f"{core_val} complete in 2026"
        return None

    long_df["completion_status"] = long_df.apply(get_completion_status, axis=1)
    long_df = long_df.drop(columns=[c for c in existing_completion_cols if c in long_df.columns])
    long_df = long_df.drop(columns=[c for c in drop_extraneous_time_cols if c in long_df.columns])

    print(f"Long format data shape: {long_df.shape}")
    print(f"Long format columns: {len(long_df.columns)}")
    print("\nNew column order:")
    print(long_df.columns.tolist())

    wb = openpyxl.load_workbook(OUTPUT_FILE)
    if "Temp" in wb.sheetnames:
        del wb["Temp"]
    if "EPI-Child-long" in wb.sheetnames:
        del wb["EPI-Child-long"]
    ws = wb.create_sheet("EPI-Child-long")

    for col_idx, col_name in enumerate(long_df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    for row_idx, row_data in enumerate(long_df.itertuples(index=False, name=None), 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(OUTPUT_FILE)
    print(f"OK: EPI-Child-long sheet created: {len(long_df)} rows")


def step_transform_pregnancy_long() -> None:
    """Transform EPI-Pregnancy to long format."""
    print_header("Step 2: Transform EPI-Pregnancy Data to Long Format")

    print("Reading EPI-Pregnancy sheet...")
    pregnancy_source = get_sheet_source_file("EPI-Pregnancy")
    required_probe_cols = {"Pregnance_code", "Td (1st) dose", "Td (2nd ) dose", "Session Point"}
    df = None
    for header_row in (0, 1):
        candidate = pd.read_excel(pregnancy_source, sheet_name="EPI-Pregnancy", header=header_row)
        if required_probe_cols.intersection(set(candidate.columns)):
            df = candidate
            print(f"Using EPI-Pregnancy header row: {header_row + 1}")
            break

    if df is None:
        raise ValueError(
            "Could not detect EPI-Pregnancy header row. Expected columns like "
            "'Pregnance_code', 'Td (1st) dose', or 'Td (2nd ) dose'."
        )
    print(f"Original data shape: {df.shape}")

    keep_columns = df.columns[:10].tolist()
    print(f"Keep columns (first 10): {keep_columns}")

    td_mappings = {
        "Td1": {
            "dose": "Td (1st) dose",
            "dose_other": "Td (1st) dose_other (Y/N)",
            "reporting_month": "Td (1st)_reporting_month",
            "source": "Td (1st) Source",
        },
        "Td2": {
            "dose": "Td (2nd ) dose",
            "dose_other": "Td (2nd) dose other (Y/N)",
            "reporting_month": "Td (2nd) dose_reporting_month",
            "source": "Td (2 nd) Source",
        },
    }

    print("\nBuilding long-form from wide sheet...")
    long_frames = []

    for dose_name, cols_map in td_mappings.items():
        required_cols = [cols_map["dose"], cols_map["reporting_month"], cols_map["source"]]
        missing = [c for c in required_cols if c not in df.columns]
        if missing:
            print(f"  Warning: missing columns for {dose_name}: {missing}")
            continue

        subset_cols = keep_columns + [
            cols_map["dose"],
            cols_map["reporting_month"],
            cols_map["source"],
        ]
        subdf = df[subset_cols].copy()

        subdf.rename(
            columns={
                cols_map["dose"]: "dose_value",
                cols_map["reporting_month"]: "reporting_month",
                cols_map["source"]: "source",
            },
            inplace=True,
        )

        subdf["Doses"] = dose_name

        subdf = subdf[
            (subdf["dose_value"].notna() & (subdf["dose_value"] != ""))
            | (subdf["reporting_month"].notna())
            | (subdf["source"].notna() & (subdf["source"] != ""))
        ]

        ordered_cols = keep_columns + ["Doses", "reporting_month", "source"]
        subdf = subdf[ordered_cols]

        long_frames.append(subdf)
        print(f"  {dose_name}: {len(subdf)} rows")

    long_df = pd.concat(long_frames, ignore_index=True)
    print(f"Long format shape: {long_df.shape}")

    # Add to workbook
    wb = openpyxl.load_workbook(OUTPUT_FILE)
    if "Td" in wb.sheetnames:
        del wb["Td"]
    ws = wb.create_sheet("Td")

    for col_idx, col_name in enumerate(long_df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    for row_idx, row_data in enumerate(long_df.itertuples(index=False, name=None), 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(OUTPUT_FILE)
    print(f"OK: Td sheet created: {len(long_df)} rows")


def step_verify_output() -> None:
    """Verify output data quality."""
    print_header("Step 3: Verify Output")

    df = pd.read_excel(OUTPUT_FILE, sheet_name="EPI-Child-long")

    print(f"Total rows: {len(df)}")
    print(f"Total columns: {len(df.columns)}")
    print(f"\nRows with vaccine_dose: {df['vaccine_dose'].notna().sum()}/{len(df)}")
    print(f"Rows with age_at_dose: {df['age_at_dose'].notna().sum()}/{len(df)}")
    print(f"Rows with reporting_month: {df['reporting_month'].notna().sum()}/{len(df)}")
    print("OK: Verification complete!")


def step_create_summary_sheet() -> None:
    """Build Summary sheet with period and clinic information."""
    print_header("Step 4: Create Summary Sheet")

    print("Reading EPI-Child-long sheet...")
    df = pd.read_excel(OUTPUT_FILE, sheet_name="EPI-Child-long")

    summary_df = df[["reporting_month", "township_name_MIMU", "Session Point"]].copy()

    print("Calculating period from reporting_month...")
    summary_df["period"] = summary_df["reporting_month"].apply(get_period)
    summary_df["township_name_MIMU"] = summary_df["township_name_MIMU"].apply(
        normalize_text
    )
    summary_df["Session Point"] = summary_df["Session Point"].apply(normalize_text)

    summary_df["Year"] = summary_df["period"].str.split("_").str[1]

    summary_df["Organization"] = "CHDN"
    summary_df["Project Name"] = "REACH-KK"
    summary_df["District (EHO)"] = ""
    summary_df["Township_EHO"] = ""

    summary_df = summary_df.rename(
        columns={"township_name_MIMU": "Twp_MIMU", "Session Point": "Clinic Name"}
    )

    final_cols = [
        "Year",
        "period",
        "Organization",
        "Project Name",
        "District (EHO)",
        "Township_EHO",
        "Twp_MIMU",
        "Clinic Name",
    ]
    summary_df = summary_df[final_cols]

    summary_df = summary_df[summary_df["period"].notna()]
    summary_df = summary_df.drop_duplicates()
    summary_df = summary_df.sort_values(["period", "Clinic Name"]).reset_index(drop=True)

    print(f"Summary sheet shape: {summary_df.shape}")

    # Add to workbook
    wb = openpyxl.load_workbook(OUTPUT_FILE)
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    ws = wb.create_sheet("Summary")

    for col_idx, col_name in enumerate(summary_df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    for row_idx, row_data in enumerate(summary_df.itertuples(index=False, name=None), 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(OUTPUT_FILE)
    print(f"OK: Summary sheet created: {len(summary_df)} rows")


def step_add_alod_columns() -> None:
    """Add ALOD (At Least One Dose) columns to summary sheet."""
    print_header("Step 5: Add ALOD Columns")

    print("Reading data...")
    child_df = pd.read_excel(OUTPUT_FILE, sheet_name="EPI-Child-long")
    summary_df = pd.read_excel(OUTPUT_FILE, sheet_name="Summary")

    print(f"EPI-Child-long shape: {child_df.shape}")
    print(f"Summary shape: {summary_df.shape}")

    # Filter for CHDN source only
    chdn_df = child_df[child_df["source"] == "CHDN"].copy()
    print(f"CHDN records: {len(chdn_df)}")

    print("\nCalculating ALOD counts...")

    chdn_df["period"] = chdn_df["reporting_month"].apply(get_period)
    chdn_df["Session Point"] = chdn_df["Session Point"].apply(normalize_text)
    chdn_df["township_name_MIMU"] = chdn_df["township_name_MIMU"].apply(normalize_text)
    chdn_df["age_first_num"] = pd.to_numeric(chdn_df["Age at first visit"], errors="coerce")

    alod_base = chdn_df[chdn_df["vaccine_dose"].notna() & chdn_df["period"].notna()].copy()

    u1_counts = (
        alod_base[
            (alod_base["age_first_num"] >= 0) & (alod_base["age_first_num"] <= 11)
        ]
        .groupby(["township_name_MIMU", "Session Point", "period"])["children_code(T/C Code)"]
        .nunique()
        .reset_index(name="ALOD-U1")
    )

    u5_counts = (
        alod_base[
            (alod_base["age_first_num"] >= 12) & (alod_base["age_first_num"] <= 59)
        ]
        .groupby(["township_name_MIMU", "Session Point", "period"])["children_code(T/C Code)"]
        .nunique()
        .reset_index(name="ALOD-U5")
    )

    gt5_counts = (
        alod_base[alod_base["age_first_num"] >= 60]
        .groupby(["township_name_MIMU", "Session Point", "period"])["children_code(T/C Code)"]
        .nunique()
        .reset_index(name="ALOD->5")
    )

    alod_counts = u1_counts.merge(u5_counts, on=["township_name_MIMU", "Session Point", "period"], how="outer").merge(
        gt5_counts, on=["township_name_MIMU", "Session Point", "period"], how="outer"
    )

    alod_counts = alod_counts.fillna(0)
    for col in ["ALOD-U1", "ALOD-U5", "ALOD->5"]:
        alod_counts[col] = alod_counts[col].astype(int)

    alod_counts = alod_counts.rename(
        columns={"township_name_MIMU": "Twp_MIMU", "Session Point": "Clinic Name"}
    )

    print(f"ALOD counts shape: {alod_counts.shape}")

    # Add ALOD columns to summary sheet
    print("\nAdding ALOD columns to summary sheet...")
    summary_df["Clinic Name"] = summary_df["Clinic Name"].apply(normalize_text)
    summary_df["Twp_MIMU"] = summary_df["Twp_MIMU"].apply(normalize_text)

    summary_df = summary_df.drop(columns=[c for c in ["ALOD-U1", "ALOD-U5", "ALOD->5"] if c in summary_df.columns])

    summary_df = summary_df.merge(
        alod_counts[["Twp_MIMU", "Clinic Name", "period", "ALOD-U1", "ALOD-U5", "ALOD->5"]],
        on=["Twp_MIMU", "Clinic Name", "period"],
        how="left",
    )

    for col in ["ALOD-U1", "ALOD-U5", "ALOD->5"]:
        summary_df[col] = pd.to_numeric(summary_df[col], errors="coerce").fillna(0).astype(int)

    print(f"Updated summary shape: {summary_df.shape}")

    # Write back to Excel
    wb = openpyxl.load_workbook(OUTPUT_FILE)
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    ws = wb.create_sheet("Summary")

    for col_idx, col_name in enumerate(summary_df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    for row_idx, row_data in enumerate(summary_df.itertuples(index=False, name=None), 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(OUTPUT_FILE)
    print("OK: ALOD columns added")
    print(f"  Total ALOD-U1: {summary_df['ALOD-U1'].sum()}")
    print(f"  Total ALOD-U5: {summary_df['ALOD-U5'].sum()}")
    print(f"  Total ALOD->5: {summary_df['ALOD->5'].sum()}")


def step_add_vaccine_columns() -> None:
    """Add vaccine-specific columns to summary sheet."""
    print_header("Step 6: Add Vaccine Columns")

    print("Reading data...")
    child_df = pd.read_excel(OUTPUT_FILE, sheet_name="EPI-Child-long")
    summary_df = pd.read_excel(OUTPUT_FILE, sheet_name="Summary")
    td_df = pd.read_excel(OUTPUT_FILE, sheet_name="Td")

    print(f"EPI-Child-long shape: {child_df.shape}")
    print(f"Summary shape: {summary_df.shape}")
    print(f"Td shape: {td_df.shape}")

    chdn_df = child_df[child_df["source"] == "CHDN"].copy()
    print(f"CHDN records: {len(chdn_df)}")

    chdn_df["period"] = chdn_df["reporting_month"].apply(get_period)
    chdn_df["age_group"] = chdn_df["age_at_dose"].apply(get_age_group)
    chdn_df["Session Point"] = chdn_df["Session Point"].apply(normalize_text)
    chdn_df["township_name_MIMU"] = chdn_df["township_name_MIMU"].apply(normalize_text)
    chdn_df["age_at_dose_num"] = pd.to_numeric(chdn_df["age_at_dose"], errors="coerce")
    chdn_df["age_first_num"] = pd.to_numeric(chdn_df["Age at first visit"], errors="coerce")

    def parse_completion(status: str):
        if pd.isna(status):
            return None
        s = str(status).strip()
        match = re.match(r"^(U1|U5|>5|1-5)\s+complete in\s+(Q[1-4])_(\d{4})$", s)
        if match:
            return f"{match.group(2)}_{match.group(3)}"
        return None

    def get_age_group_first(age):
        if pd.isna(age):
            return None
        if age <= 11:
            return "U1"
        if age <= 59:
            return "U5"
        return ">5"

    chdn_df["completion_period"] = chdn_df["completion_status"].apply(parse_completion)
    chdn_df["completion_age"] = chdn_df["age_first_num"].apply(get_age_group_first)

    vaccines = ["BCG", "OPV1", "OPV2", "OPV3", "Penta1", "Penta2", "Penta3", "MMR1", "MMR2", "JE1", "IPV"]
    age_groups = ["U1", "U5", ">5"]

    print("\nPreparing vaccine-specific counts by clinic, period, and age group...")
    print(f"Unique vaccines in data: {sorted(chdn_df['vaccine_dose'].dropna().unique())}")

    summary_df["Clinic Name"] = summary_df["Clinic Name"].apply(normalize_text)
    summary_df["Twp_MIMU"] = summary_df["Twp_MIMU"].apply(normalize_text)
    summary_df = summary_df.drop_duplicates(subset=["period", "Twp_MIMU", "Clinic Name"]).reset_index(drop=True)
    clinic_period_combos = summary_df[["Twp_MIMU", "Clinic Name", "period"]].drop_duplicates()

    td_chdn_temp = td_df[td_df["source"] == "CHDN"].copy()
    td_chdn_temp["period"] = td_chdn_temp["reporting_month"].apply(get_period)
    td_chdn_temp["Clinic Name"] = td_chdn_temp["Session Point"].apply(normalize_text)
    td_chdn_temp["Twp_MIMU"] = td_chdn_temp["township_name_MIMU"].apply(normalize_text)
    td_chdn_temp = td_chdn_temp[td_chdn_temp["period"].notna()]
    td_pairs = td_chdn_temp[["Twp_MIMU", "Clinic Name", "period"]].drop_duplicates()
    clinic_period_combos = pd.concat([clinic_period_combos, td_pairs]).drop_duplicates().reset_index(drop=True)
    print(f"\nClinic-period combinations (including Td): {len(clinic_period_combos)}")

    def build_summary_key(df, township_col, clinic_col, period_col="period"):
        return (
            df[township_col].apply(normalize_text)
            + "|"
            + df[clinic_col].apply(normalize_text)
            + "|"
            + df[period_col].astype(str)
        )

    def get_alod_age_group(age):
        if pd.isna(age):
            return None
        try:
            age_val = int(float(age))
        except Exception:
            return None
        if age_val <= 11:
            return "U1"
        if age_val <= 59:
            return "U5"
        return ">5"

    vaccine_col_order = [
        "BCG_U1", "BCG_U5", "BCG_>5",
        "OPV1_U1", "OPV1_U5", "OPV1_>5",
        "OPV2_U1", "OPV2_U5", "OPV2_>5",
        "OPV3_U1", "OPV3_U5", "OPV3_>5",
        "Penta1_U1", "Penta1_U5", "Penta1_>5",
        "Penta2_U1", "Penta2_U5", "Penta2_>5",
        "Penta3_U1", "Penta3_U5", "Penta3_>5",
        "MMR1_U1", "MMR1_U5", "MMR1_>5",
        "MMR2_U1", "MMR2_U5", "MMR2_>5",
        "JE_U1", "JE_U5", "JE_>5",
        "IPV_U1", "IPV_U5", "IPV_>5",
    ]

    vaccine_columns_to_add = {}
    print("\nCreating vaccine-age matrix...")
    u1_overrides = {
        "Penta1": (2, 11),
        "Penta3": (4, 11),
        "MMR1": (9, 11),
        "MMR2": (10, 11),
    }

    for vaccine in vaccines:
        for age_group in age_groups:
            display_vaccine = "JE" if vaccine == "JE1" else vaccine
            col_name = f"{display_vaccine}_{age_group}"
            base = chdn_df[chdn_df["vaccine_dose"] == vaccine].copy()

            if age_group == "U1":
                min_age, max_age = u1_overrides.get(display_vaccine, (0, 11))
                subset_df = base[(base["age_at_dose_num"] >= min_age) & (base["age_at_dose_num"] <= max_age)]
            elif age_group == "U5":
                subset_df = base[(base["age_at_dose_num"] >= 12) & (base["age_at_dose_num"] <= 59)]
            else:
                subset_df = base[(base["age_at_dose_num"] >= 60)]

            if subset_df.empty:
                print(f"  {col_name}: No data (filled with 0)")
                col_data = pd.DataFrame({
                    "Twp_MIMU": clinic_period_combos["Twp_MIMU"],
                    "Clinic Name": clinic_period_combos["Clinic Name"],
                    "period": clinic_period_combos["period"],
                    "count": 0,
                })
            else:
                counts = subset_df.groupby(["township_name_MIMU", "Session Point", "period"])["children_code(T/C Code)"] \
                    .nunique().reset_index(name="count")
                counts = counts.rename(columns={"township_name_MIMU": "Twp_MIMU", "Session Point": "Clinic Name"})
                counts["Twp_MIMU"] = counts["Twp_MIMU"].apply(normalize_text)
                counts["Clinic Name"] = counts["Clinic Name"].apply(normalize_text)
                col_data = clinic_period_combos.merge(counts, on=["Twp_MIMU", "Clinic Name", "period"], how="left")
                col_data["count"] = col_data["count"].fillna(0).astype(int)

            col_data["key"] = build_summary_key(col_data, "Twp_MIMU", "Clinic Name")
            vaccine_columns_to_add[col_name] = dict(zip(col_data["key"], col_data["count"]))

    print(f"\nCreated {len(vaccine_columns_to_add)} vaccine-age columns")
    print("\nCreating completion-age matrix (CD columns)...")
    cd_columns_to_add = {}
    for comp_age in ["U1", "U5", ">5"]:
        col_name = f"CD_{comp_age}"
        comp_df = chdn_df[(chdn_df["completion_age"] == comp_age) & chdn_df["completion_period"].notna()].copy()
        if comp_df.empty:
            print(f"  {col_name}: No data (filled with 0)")
            col_data = pd.DataFrame({
                "Twp_MIMU": clinic_period_combos["Twp_MIMU"],
                "Clinic Name": clinic_period_combos["Clinic Name"],
                "period": clinic_period_combos["period"],
                "count": 0,
            })
        else:
            counts = comp_df.groupby(["township_name_MIMU", "Session Point", "completion_period"])["children_code(T/C Code)"] \
                .nunique().reset_index(name="count")
            counts = counts.rename(columns={
                "township_name_MIMU": "Twp_MIMU",
                "Session Point": "Clinic Name",
                "completion_period": "period",
            })
            counts["Twp_MIMU"] = counts["Twp_MIMU"].apply(normalize_text)
            counts["Clinic Name"] = counts["Clinic Name"].apply(normalize_text)
            col_data = clinic_period_combos.merge(counts, on=["Twp_MIMU", "Clinic Name", "period"], how="left")
            col_data["count"] = col_data["count"].fillna(0).astype(int)
        col_data["key"] = build_summary_key(col_data, "Twp_MIMU", "Clinic Name")
        cd_columns_to_add[col_name] = dict(zip(col_data["key"], col_data["count"]))

    print("\nCreating Td dose columns from Td sheet...")
    td_chdn = td_df[td_df["source"] == "CHDN"].copy()
    td_chdn["period"] = td_chdn["reporting_month"].apply(get_period)
    td_chdn = td_chdn[td_chdn["period"].notna()]
    td_chdn["Session Point"] = td_chdn["Session Point"].apply(normalize_text)
    td_chdn["township_name_MIMU"] = td_chdn["township_name_MIMU"].apply(normalize_text)

    td_columns_to_add = {}
    for dose in ["Td1", "Td2"]:
        dose_df = td_chdn[td_chdn["Doses"] == dose].copy()
        if dose_df.empty:
            print(f"  {dose}: No data (filled with 0)")
            col_data = pd.DataFrame({
                "Twp_MIMU": clinic_period_combos["Twp_MIMU"],
                "Clinic Name": clinic_period_combos["Clinic Name"],
                "period": clinic_period_combos["period"],
                "count": 0,
            })
        else:
            counts = dose_df.groupby(["township_name_MIMU", "Session Point", "period"])["Pregnance_code"] \
                .nunique().reset_index(name="count")
            counts = counts.rename(columns={"township_name_MIMU": "Twp_MIMU", "Session Point": "Clinic Name"})
            counts["Twp_MIMU"] = counts["Twp_MIMU"].apply(normalize_text)
            counts["Clinic Name"] = counts["Clinic Name"].apply(normalize_text)
            col_data = clinic_period_combos.merge(counts, on=["Twp_MIMU", "Clinic Name", "period"], how="left")
            col_data["count"] = col_data["count"].fillna(0).astype(int)
        col_data["key"] = build_summary_key(col_data, "Twp_MIMU", "Clinic Name")
        td_columns_to_add[dose] = dict(zip(col_data["key"], col_data["count"]))

    print("\nCreating Td at-least-one column...")
    if td_chdn.empty:
        print("  Td At least one dose: No data (filled with 0)")
        td_any_col = pd.DataFrame({
            "Twp_MIMU": clinic_period_combos["Twp_MIMU"],
            "Clinic Name": clinic_period_combos["Clinic Name"],
            "period": clinic_period_combos["period"],
            "count": 0,
        })
    else:
        any_counts = td_chdn.groupby(["township_name_MIMU", "Session Point", "period"])["Pregnance_code"] \
            .nunique().reset_index(name="count")
        any_counts = any_counts.rename(columns={"township_name_MIMU": "Twp_MIMU", "Session Point": "Clinic Name"})
        any_counts["Twp_MIMU"] = any_counts["Twp_MIMU"].apply(normalize_text)
        any_counts["Clinic Name"] = any_counts["Clinic Name"].apply(normalize_text)
        td_any_col = clinic_period_combos.merge(any_counts, on=["Twp_MIMU", "Clinic Name", "period"], how="left")
        td_any_col["count"] = td_any_col["count"].fillna(0).astype(int)
    td_any_col["key"] = build_summary_key(td_any_col, "Twp_MIMU", "Clinic Name")
    td_columns_to_add["Td At least one dose"] = dict(zip(td_any_col["key"], td_any_col["count"]))

    print("\nAdding new clinic-period rows from Td data...")
    expanded_pairs = clinic_period_combos.copy()
    summary_pairs = summary_df[["Twp_MIMU", "Clinic Name", "period"]].copy()
    new_pairs = expanded_pairs.merge(summary_pairs, on=["Twp_MIMU", "Clinic Name", "period"], how="left", indicator=True)
    new_pairs = new_pairs[new_pairs["_merge"] == "left_only"][["Twp_MIMU", "Clinic Name", "period"]].drop_duplicates()

    if len(new_pairs) > 0:
        print(f"Found {len(new_pairs)} new clinic-period rows from Td data...")
        new_rows = new_pairs.copy()
        new_rows["Organization"] = "CHDN"
        new_rows["Project Name"] = "REACH-KK"
        new_rows["District (EHO)"] = ""
        new_rows["Township_EHO"] = ""
        new_rows["Year"] = new_rows["period"].astype(str).str.split("_").str[1]
        for col in vaccine_col_order + ["ALOD-U1", "ALOD-U5", "ALOD->5", "CD_U1", "CD_U5", "CD_>5", "Td1", "Td2", "Td At least one dose"]:
            new_rows[col] = 0
        for col in summary_df.columns:
            if col not in new_rows.columns:
                new_rows[col] = ""
        new_rows = new_rows[summary_df.columns]
        summary_df = pd.concat([summary_df, new_rows], ignore_index=True)
        print(f"Added {len(new_pairs)} new clinic-period rows")

    print("\nMapping vaccine-age columns to summary sheet...")
    summary_df["Clinic Name"] = summary_df["Clinic Name"].apply(normalize_text)
    summary_df["Twp_MIMU"] = summary_df["Twp_MIMU"].apply(normalize_text)
    summary_df = summary_df.drop_duplicates(subset=["period", "Twp_MIMU", "Clinic Name"]).reset_index(drop=True)
    summary_df["key"] = build_summary_key(summary_df, "Twp_MIMU", "Clinic Name")

    for col_name, mapping in vaccine_columns_to_add.items():
        summary_df[col_name] = summary_df["key"].map(mapping).fillna(0).astype(int)
    for col_name, mapping in cd_columns_to_add.items():
        summary_df[col_name] = summary_df["key"].map(mapping).fillna(0).astype(int)
    for col_name, mapping in td_columns_to_add.items():
        summary_df[col_name] = summary_df["key"].map(mapping).fillna(0).astype(int)

    alod_cols = ["ALOD-U1", "ALOD-U5", "ALOD->5"]
    alod_missing = any(col not in summary_df.columns for col in alod_cols)
    alod_zero_only = False
    if not alod_missing:
        alod_total = pd.to_numeric(summary_df[alod_cols].stack(), errors="coerce").fillna(0).sum()
        alod_zero_only = alod_total == 0 and not chdn_df.empty

    if alod_missing or alod_zero_only:
        reason = "missing from Summary" if alod_missing else "present but all zero"
        print(f"\nALOD columns are {reason}. Recomputing them from EPI-Child-long data...")
        alod_df = chdn_df[["township_name_MIMU", "Session Point", "period", "age_first_num", "children_code(T/C Code)", "vaccine_dose"]].copy()
        alod_df = alod_df[alod_df["vaccine_dose"].notna() & alod_df["period"].notna()].copy()
        alod_df["age_group"] = alod_df["age_first_num"].apply(get_alod_age_group)
        alod_counts = (
            alod_df.dropna(subset=["age_group"])
            .groupby(["township_name_MIMU", "Session Point", "period", "age_group"])["children_code(T/C Code)"]
            .nunique()
            .unstack(fill_value=0)
            .reset_index()
            .rename(columns={
                "township_name_MIMU": "Twp_MIMU",
                "Session Point": "Clinic Name",
                "U1": "ALOD-U1",
                "U5": "ALOD-U5",
                ">5": "ALOD->5",
            })
        )
        alod_counts["Twp_MIMU"] = alod_counts["Twp_MIMU"].apply(normalize_text)
        alod_counts["Clinic Name"] = alod_counts["Clinic Name"].apply(normalize_text)
        for col in alod_cols:
            if col not in alod_counts.columns:
                alod_counts[col] = 0

        summary_df = summary_df.drop(columns=[col for col in alod_cols if col in summary_df.columns])
        summary_df = summary_df.merge(alod_counts[["Twp_MIMU", "Clinic Name", "period"] + alod_cols], on=["Twp_MIMU", "Clinic Name", "period"], how="left")
        for col in alod_cols:
            summary_df[col] = pd.to_numeric(summary_df[col], errors="coerce").fillna(0).astype(int)

    summary_df = summary_df.drop(columns=["key"])
    base_cols = ["Year", "period", "Organization", "Project Name", "District (EHO)", "Township_EHO", "Twp_MIMU", "Clinic Name", "ALOD-U1", "ALOD-U5", "ALOD->5"]
    final_col_order = base_cols + vaccine_col_order + ["CD_U1", "CD_U5", "CD_>5", "Td1", "Td2", "Td At least one dose"]

    for col in final_col_order:
        if col not in summary_df.columns:
            summary_df[col] = 0 if col not in base_cols[:8] else ""

    summary_df = summary_df[final_col_order]
    print(f"Updated summary shape: {summary_df.shape}")
    print(f"Total columns: {len(summary_df.columns)}")

    wb = openpyxl.load_workbook(OUTPUT_FILE)
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    ws = wb.create_sheet("Summary")
    
    for col_idx, col_name in enumerate(summary_df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    for row_idx, row_data in enumerate(summary_df.itertuples(index=False, name=None), 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(OUTPUT_FILE)
    print(f"Successfully updated 'Summary' sheet with vaccine-age columns")
    print(f"\nTotal columns in Summary sheet: {len(summary_df.columns)}")
    print(f"Total rows: {len(summary_df)}")


def step_add_cumulative_sheet() -> None:
    """Build yearly_cumulative sheet."""
    print_header("Step 7: Add Cumulative Sheet")

    print("Reading data...")
    child_df = pd.read_excel(OUTPUT_FILE, sheet_name="EPI-Child-long")
    td_df = pd.read_excel(OUTPUT_FILE, sheet_name="Td")

    chdn_df = child_df[child_df["source"] == "CHDN"].copy()
    print(f"CHDN records: {len(chdn_df)}")

    def get_year(date):
        if pd.isna(date):
            return None
        return str(date.year)

    chdn_df["year"] = chdn_df["reporting_month"].apply(get_year)
    chdn_df["age_group"] = chdn_df["Age at first visit"].apply(get_age_group)
    chdn_df["Session Point"] = chdn_df["Session Point"].astype(str).str.strip()

    missing_age = chdn_df[chdn_df["age_group"].isna()]
    print(f"\nRecords with missing/invalid age data: {len(missing_age)}")
    print(f"Unique children_code with missing age: {missing_age['children_code(T/C Code)'].nunique()}")

    cumulative_df = chdn_df[["year", "township_name_MIMU", "Session Point", "children_code(T/C Code)", "age_group"]].copy()
    cumulative_df = cumulative_df[cumulative_df["year"].notna()].copy()
    unique_combos = cumulative_df[["year", "township_name_MIMU", "Session Point"]].drop_duplicates()
    print(f"\nUnique year-clinic combinations: {len(unique_combos)}")

    print("\nCalculating ALOD counts by clinic, year, and age group...")
    alod_data = []
    for _, row in unique_combos.iterrows():
        year = row["year"]
        township = row["township_name_MIMU"]
        clinic = row["Session Point"]
        clinic_data = cumulative_df[(cumulative_df["year"] == year) & (cumulative_df["Session Point"] == clinic)]
        alod_u1 = len(clinic_data[(clinic_data["age_group"] == "U1")]["children_code(T/C Code)"].unique())
        alod_u5 = len(clinic_data[(clinic_data["age_group"] == "U5")]["children_code(T/C Code)"].unique())
        alod_gt5 = len(clinic_data[(clinic_data["age_group"] == ">5")]["children_code(T/C Code)"].unique())
        alod_data.append({
            "year": year,
            "township_name_MIMU": township,
            "Session Point": clinic,
            "ALOD-U1": alod_u1,
            "ALOD-U5": alod_u5,
            "ALOD->5": alod_gt5,
        })

    alod_df = pd.DataFrame(alod_data)
    print(f"ALOD calculations shape: {alod_df.shape}")
    print(alod_df.head(10))

    print("\nCalculating Td at least one dose by clinic and year...")
    td_chdn = td_df[td_df["source"] == "CHDN"].copy()
    print(f"Total Td CHDN records: {len(td_chdn)}")
    print(f"Unique Pregnance_code in all Td CHDN data: {td_chdn['Pregnance_code'].nunique()}")
    td_chdn["year"] = td_chdn["reporting_month"].apply(get_year)
    td_chdn["Session Point"] = td_chdn["Session Point"].astype(str).str.strip()
    null_reporting = td_chdn[td_chdn["year"].isna()]
    print(f"Td records with null reporting_month: {len(null_reporting)}")
    print(f"Unique Pregnance_code in records with null reporting_month: {null_reporting['Pregnance_code'].nunique()}")
    td_chdn = td_chdn[td_chdn["year"].notna()]
    print(f"Td CHDN records after filtering null reporting_month: {len(td_chdn)}")
    print(f"Unique Pregnance_code after filtering: {td_chdn['Pregnance_code'].nunique()}")
    td_counts = td_chdn.groupby(["Session Point", "year"])["Pregnance_code"].nunique().reset_index(name="Td At least one dose")
    td_counts = td_counts.rename(columns={"Session Point": "clinic_temp"})
    print(f"Td counts shape: {td_counts.shape}")

    alod_df["clinic_temp"] = alod_df["Session Point"]
    alod_df = alod_df.merge(td_counts, on=["clinic_temp", "year"], how="left")
    alod_df["Td At least one dose"] = alod_df["Td At least one dose"].fillna(0).astype(int)
    alod_df = alod_df.drop(columns=["clinic_temp"])

    summary_cumulative = alod_df[["year", "township_name_MIMU", "Session Point"]].copy()
    summary_cumulative["Organization"] = "CHDN"
    summary_cumulative["Project Name"] = "REACH-KK"
    summary_cumulative["District (EHO)"] = ""
    summary_cumulative["Township_EHO"] = ""
    summary_cumulative = summary_cumulative.rename(columns={
        "year": "period",
        "township_name_MIMU": "Twp_MIMU",
        "Session Point": "Clinic Name",
    })
    summary_cumulative["ALOD-U1"] = alod_df["ALOD-U1"].values
    summary_cumulative["ALOD-U5"] = alod_df["ALOD-U5"].values
    summary_cumulative["ALOD->5"] = alod_df["ALOD->5"].values
    summary_cumulative["Td At least one dose"] = alod_df["Td At least one dose"].values

    final_cols = ["Organization", "period", "Project Name", "District (EHO)", "Township_EHO", "Twp_MIMU", "Clinic Name", "ALOD-U1", "ALOD-U5", "ALOD->5", "Td At least one dose"]
    summary_cumulative = summary_cumulative[final_cols]
    summary_cumulative = summary_cumulative.drop_duplicates()
    summary_cumulative = summary_cumulative.sort_values(["period", "Clinic Name"]).reset_index(drop=True)
    print(f"\nCumulative summary shape: {summary_cumulative.shape}")
    print(f"Unique periods: {sorted(summary_cumulative['period'].unique())}")
    print(f"Unique clinics: {summary_cumulative['Clinic Name'].nunique()}")
    print("\nSample data:")
    print(summary_cumulative.head(10))

    wb = openpyxl.load_workbook(OUTPUT_FILE)
    if "yearly_cumulative" in wb.sheetnames:
        del wb["yearly_cumulative"]
    ws = wb.create_sheet("yearly_cumulative")

    for col_idx, col_name in enumerate(summary_cumulative.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    for row_idx, row_data in enumerate(summary_cumulative.itertuples(index=False, name=None), 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    print(f"Successfully added 'yearly_cumulative' sheet")
    print("\nYearly Cumulative Summary:")
    print(f"  Total rows: {len(summary_cumulative)}")
    print(f"  Periods (years): {sorted(summary_cumulative['period'].unique())}")
    print(f"  Total ALOD-U1: {summary_cumulative['ALOD-U1'].sum()}")
    print(f"  Total ALOD-U5: {summary_cumulative['ALOD-U5'].sum()}")
    print(f"  Total ALOD->5: {summary_cumulative['ALOD->5'].sum()}")
    print(f"  Total Td at least one dose: {summary_cumulative['Td At least one dose'].sum()}")

    print("\n" + "=" * 40)
    print("Building overall cumulative sheet...")
    print("=" * 40)
    print("\nCalculating unique children_code across all years by clinic and age group...")

    overall_data = []
    unique_clinics = chdn_df[["township_name_MIMU", "Session Point"]].drop_duplicates()
    for _, row in unique_clinics.iterrows():
        township = row["township_name_MIMU"]
        clinic = row["Session Point"]
        clinic_data = chdn_df[chdn_df["Session Point"] == clinic]
        alod_u1 = len(clinic_data[(clinic_data["age_group"] == "U1")]["children_code(T/C Code)"].unique())
        alod_u5 = len(clinic_data[(clinic_data["age_group"] == "U5")]["children_code(T/C Code)"].unique())
        alod_gt5 = len(clinic_data[(clinic_data["age_group"] == ">5")]["children_code(T/C Code)"].unique())
        td_clinic = td_chdn[td_chdn["Session Point"] == clinic]
        td_dose = len(td_clinic["Pregnance_code"].unique())
        overall_data.append({
            "Organization": "CHDN",
            "period": "overall",
            "Project Name": "REACH-KK",
            "District (EHO)": "",
            "Township_EHO": "",
            "Twp_MIMU": township,
            "Clinic Name": clinic,
            "ALOD-U1": alod_u1,
            "ALOD-U5": alod_u5,
            "ALOD->5": alod_gt5,
            "Td At least one dose": td_dose,
        })

    summary_overall = pd.DataFrame(overall_data)
    summary_overall = summary_overall.sort_values(["Clinic Name"]).reset_index(drop=True)
    print(f"Overall cumulative shape: {summary_overall.shape}")
    print("\nSample data:")
    print(summary_overall.head(10))

    if "cumulative" in wb.sheetnames:
        del wb["cumulative"]
        print(f"  Removed existing 'cumulative' sheet")
    ws = wb.create_sheet("cumulative")
    for col_idx, col_name in enumerate(summary_overall.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    for row_idx, row_data in enumerate(summary_overall.itertuples(index=False, name=None), 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(OUTPUT_FILE)
    print(f"Successfully added 'cumulative' sheet")
    print("\nOverall Cumulative Summary:")
    print(f"  Total rows: {len(summary_overall)}")
    print(f"  Total ALOD-U1: {summary_overall['ALOD-U1'].sum()}")
    print(f"  Total ALOD-U5: {summary_overall['ALOD-U5'].sum()}")
    print(f"  Total ALOD->5: {summary_overall['ALOD->5'].sum()}")
    print(f"  Total Td at least one dose: {summary_overall['Td At least one dose'].sum()}")


def step_add_cumulative() -> None:
    """Add cumulative sheet (similar to yearly but with additional data)."""
    print_header("Step 7b: Add Cumulative Sheet")
    print("Cumulative sheet already built inside step_add_cumulative_sheet().")


def step_add_multi_year_indicators() -> None:
    """Add multi-year indicators sheet."""
    print_header("Step 8: Add Multi-Year Indicators Sheet")

    print("Reading data...")
    child_df = pd.read_excel(OUTPUT_FILE, sheet_name="EPI-Child-long")
    td_df = pd.read_excel(OUTPUT_FILE, sheet_name="Td")
    indicators_source = get_sheet_source_file("indicators")
    try:
        template_df = pd.read_excel(indicators_source, sheet_name="indicators", engine="openpyxl")
    except ValueError:
        template_df = pd.read_excel(indicators_source, sheet_name="2025_indicators", engine="openpyxl")

    print(f"EPI-Child-long shape: {child_df.shape}")
    print(f"Td shape: {td_df.shape}")
    print(f"Template shape: {template_df.shape}")

    chdn_df = child_df[child_df["source"] == "CHDN"].copy()
    chdn_df["year"] = pd.to_datetime(chdn_df["reporting_month"], errors="coerce").dt.year

    td_chdn = td_df[td_df["source"] == "CHDN"].copy()
    td_chdn["year"] = pd.to_datetime(td_chdn["reporting_month"], errors="coerce").dt.year

    def get_quarter(date):
        if pd.isna(date):
            return None
        d = pd.to_datetime(date)
        y = d.year
        q1_start = pd.Timestamp(year=y - 1, month=12, day=21)
        q1_end = pd.Timestamp(year=y, month=3, day=20)
        if q1_start <= d <= q1_end:
            return "Q1"
        q2_start = pd.Timestamp(year=y, month=3, day=21)
        q2_end = pd.Timestamp(year=y, month=6, day=20)
        if q2_start <= d <= q2_end:
            return "Q2"
        q3_start = pd.Timestamp(year=y, month=6, day=21)
        q3_end = pd.Timestamp(year=y, month=9, day=20)
        if q3_start <= d <= q3_end:
            return "Q3"
        q4_start = pd.Timestamp(year=y, month=9, day=21)
        q4_end = pd.Timestamp(year=y, month=12, day=20)
        if q4_start <= d <= q4_end:
            return "Q4"
        return "Q1"

    chdn_df["quarter"] = chdn_df["reporting_month"].apply(get_quarter)
    chdn_df["age_at_dose_num"] = pd.to_numeric(chdn_df["age_at_dose"], errors="coerce")
    chdn_df["age_first_num"] = pd.to_numeric(chdn_df["Age at first visit"], errors="coerce")
    td_chdn["quarter"] = td_chdn["reporting_month"].apply(get_quarter)

    def parse_completion(status: str):
        if pd.isna(status):
            return None, None, None
        s = str(status).strip()
        match = re.match(r"^(U1|U5|>5|1-5)\s+complete in\s+(Q[1-4])_(\d{4})$", s)
        if match:
            age = match.group(1)
            age = "U5" if age == "1-5" else age
            quarter = match.group(2)
            year = int(match.group(3))
            return age, quarter, year
        return None, None, None

    indicators_df = template_df.copy()
    print("\nTemplate indicators:")
    print(list(indicators_df["indicator"].unique()))
    print(f"Total rows: {len(indicators_df)}")

    years_to_process = [2024, 2025, 2026]
    for year in years_to_process:
        print(f"\n{'=' * 60}")
        print(f"Processing year {year}")
        print(f"{'=' * 60}")

        chdn_year = chdn_df[chdn_df["year"] == year].copy()
        td_chdn_year = td_chdn[(td_chdn["year"] == year) & td_chdn["quarter"].notna()].copy()

        print(f"CHDN records for {year}: {len(chdn_year)}")
        print(f"CHDN Td records for {year}: {len(td_chdn_year)}")

        if len(chdn_year) == 0 and len(td_chdn_year) == 0:
            print(f"  No data for {year}, skipping...")
            continue

        if len(chdn_year) > 0:
            chdn_year[["completion_age", "completion_quarter", "completion_year"]] = chdn_year["completion_status"].apply(
                lambda x: pd.Series(parse_completion(x))
            )

        quarters = ["Q1", "Q2", "Q3", "Q4"]

        penta3_u1_indices = indicators_df[(indicators_df["Period"] == year) & (indicators_df["indicator"] == "Penta3 under 1-yr-old")].index
        if len(penta3_u1_indices) > 0:
            penta3_u1_idx = penta3_u1_indices[0]
            penta3_df = chdn_year[chdn_year["vaccine_dose"] == "Penta3"].copy()
            if len(penta3_df) > 0:
                for quarter in quarters:
                    quarter_data = penta3_df[penta3_df["quarter"] == quarter]
                    u1_data = quarter_data[(quarter_data["age_at_dose_num"] >= 4) & (quarter_data["age_at_dose_num"] <= 11)]
                    u1_male = len(u1_data[u1_data["Gender"] == "Male"]["children_code(T/C Code)"].unique())
                    u1_female = len(u1_data[u1_data["Gender"] == "Female"]["children_code(T/C Code)"].unique())
                    indicators_df.loc[penta3_u1_idx, f"{quarter} U1 Male"] = u1_male
                    indicators_df.loc[penta3_u1_idx, f"{quarter} U1 Female"] = u1_female

        penta3_u5_indices = indicators_df[(indicators_df["Period"] == year) & (indicators_df["indicator"] == "Penta3 under 5-yr-old")].index
        if len(penta3_u5_indices) > 0:
            penta3_u5_idx = penta3_u5_indices[0]
            penta3_df = chdn_year[chdn_year["vaccine_dose"] == "Penta3"].copy()
            if len(penta3_df) > 0:
                for quarter in quarters:
                    quarter_data = penta3_df[penta3_df["quarter"] == quarter]
                    u1_data = quarter_data[(quarter_data["age_at_dose_num"] >= 4) & (quarter_data["age_at_dose_num"] <= 11)]
                    u1_male = len(u1_data[u1_data["Gender"] == "Male"]["children_code(T/C Code)"].unique())
                    u1_female = len(u1_data[u1_data["Gender"] == "Female"]["children_code(T/C Code)"].unique())
                    u5_data = quarter_data[(quarter_data["age_at_dose_num"] >= 12) & (quarter_data["age_at_dose_num"] <= 59)]
                    u5_male = len(u5_data[u5_data["Gender"] == "Male"]["children_code(T/C Code)"].unique())
                    u5_female = len(u5_data[u5_data["Gender"] == "Female"]["children_code(T/C Code)"].unique())
                    indicators_df.loc[penta3_u5_idx, f"{quarter} U1 Male"] = u1_male
                    indicators_df.loc[penta3_u5_idx, f"{quarter} U1 Female"] = u1_female
                    indicators_df.loc[penta3_u5_idx, f"{quarter} 1-5 Male "] = u5_male
                    indicators_df.loc[penta3_u5_idx, f"{quarter} 1-5 Female"] = u5_female

        mmr1_u1_indices = indicators_df[(indicators_df["Period"] == year) & (indicators_df["indicator"] == "MMR1 under 1-yr-old")].index
        if len(mmr1_u1_indices) > 0:
            mmr1_u1_idx = mmr1_u1_indices[0]
            mmr1_df = chdn_year[chdn_year["vaccine_dose"] == "MMR1"].copy()
            if len(mmr1_df) > 0:
                for quarter in quarters:
                    quarter_data = mmr1_df[mmr1_df["quarter"] == quarter]
                    u1_data = quarter_data[(quarter_data["age_at_dose_num"] >= 9) & (quarter_data["age_at_dose_num"] <= 11)]
                    u1_male = len(u1_data[u1_data["Gender"] == "Male"]["children_code(T/C Code)"].unique())
                    u1_female = len(u1_data[u1_data["Gender"] == "Female"]["children_code(T/C Code)"].unique())
                    indicators_df.loc[mmr1_u1_idx, f"{quarter} U1 Male"] = u1_male
                    indicators_df.loc[mmr1_u1_idx, f"{quarter} U1 Female"] = u1_female

        mmr1_u5_indices = indicators_df[(indicators_df["Period"] == year) & (indicators_df["indicator"] == "MMR1 under 5-yr-old")].index
        if len(mmr1_u5_indices) > 0:
            mmr1_u5_idx = mmr1_u5_indices[0]
            mmr1_df = chdn_year[chdn_year["vaccine_dose"] == "MMR1"].copy()
            if len(mmr1_df) > 0:
                for quarter in quarters:
                    quarter_data = mmr1_df[mmr1_df["quarter"] == quarter]
                    u1_data = quarter_data[(quarter_data["age_at_dose_num"] >= 9) & (quarter_data["age_at_dose_num"] <= 11)]
                    u1_male = len(u1_data[u1_data["Gender"] == "Male"]["children_code(T/C Code)"].unique())
                    u1_female = len(u1_data[u1_data["Gender"] == "Female"]["children_code(T/C Code)"].unique())
                    u5_data = quarter_data[(quarter_data["age_at_dose_num"] >= 12) & (quarter_data["age_at_dose_num"] <= 59)]
                    u5_male = len(u5_data[u5_data["Gender"] == "Male"]["children_code(T/C Code)"].unique())
                    u5_female = len(u5_data[u5_data["Gender"] == "Female"]["children_code(T/C Code)"].unique())
                    indicators_df.loc[mmr1_u5_idx, f"{quarter} U1 Male"] = u1_male
                    indicators_df.loc[mmr1_u5_idx, f"{quarter} U1 Female"] = u1_female
                    indicators_df.loc[mmr1_u5_idx, f"{quarter} 1-5 Male "] = u5_male
                    indicators_df.loc[mmr1_u5_idx, f"{quarter} 1-5 Female"] = u5_female

        mmr2_u5_indices = indicators_df[(indicators_df["Period"] == year) & (indicators_df["indicator"] == "MMR2 under 5-yr-old")].index
        if len(mmr2_u5_indices) > 0:
            mmr2_u5_idx = mmr2_u5_indices[0]
            mmr2_df = chdn_year[chdn_year["vaccine_dose"] == "MMR2"].copy()
            if len(mmr2_df) > 0:
                for quarter in quarters:
                    quarter_data = mmr2_df[mmr2_df["quarter"] == quarter]
                    u1_data = quarter_data[(quarter_data["age_at_dose_num"] >= 10) & (quarter_data["age_at_dose_num"] <= 11)]
                    u1_male = len(u1_data[u1_data["Gender"] == "Male"]["children_code(T/C Code)"].unique())
                    u1_female = len(u1_data[u1_data["Gender"] == "Female"]["children_code(T/C Code)"].unique())
                    u5_data = quarter_data[(quarter_data["age_at_dose_num"] >= 12) & (quarter_data["age_at_dose_num"] <= 59)]
                    u5_male = len(u5_data[u5_data["Gender"] == "Male"]["children_code(T/C Code)"].unique())
                    u5_female = len(u5_data[u5_data["Gender"] == "Female"]["children_code(T/C Code)"].unique())
                    indicators_df.loc[mmr2_u5_idx, f"{quarter} U1 Male"] = u1_male
                    indicators_df.loc[mmr2_u5_idx, f"{quarter} U1 Female"] = u1_female
                    indicators_df.loc[mmr2_u5_idx, f"{quarter} 1-5 Male "] = u5_male
                    indicators_df.loc[mmr2_u5_idx, f"{quarter} 1-5 Female"] = u5_female

        penta1_u5_indices = indicators_df[(indicators_df["Period"] == year) & (indicators_df["indicator"] == "Penta1 under 5-yr-old")].index
        if len(penta1_u5_indices) > 0:
            penta1_u5_idx = penta1_u5_indices[0]
            penta1_df = chdn_year[chdn_year["vaccine_dose"] == "Penta1"].copy()
            if len(penta1_df) > 0:
                for quarter in quarters:
                    quarter_data = penta1_df[penta1_df["quarter"] == quarter]
                    u1_data = quarter_data[(quarter_data["age_at_dose_num"] >= 2) & (quarter_data["age_at_dose_num"] <= 11)]
                    u1_male = len(u1_data[u1_data["Gender"] == "Male"]["children_code(T/C Code)"].unique())
                    u1_female = len(u1_data[u1_data["Gender"] == "Female"]["children_code(T/C Code)"].unique())
                    u5_data = quarter_data[(quarter_data["age_at_dose_num"] >= 12) & (quarter_data["age_at_dose_num"] <= 59)]
                    u5_male = len(u5_data[u5_data["Gender"] == "Male"]["children_code(T/C Code)"].unique())
                    u5_female = len(u5_data[u5_data["Gender"] == "Female"]["children_code(T/C Code)"].unique())
                    indicators_df.loc[penta1_u5_idx, f"{quarter} U1 Male"] = u1_male
                    indicators_df.loc[penta1_u5_idx, f"{quarter} U1 Female"] = u1_female
                    indicators_df.loc[penta1_u5_idx, f"{quarter} 1-5 Male "] = u5_male
                    indicators_df.loc[penta1_u5_idx, f"{quarter} 1-5 Female"] = u5_female

        atleast_indices = indicators_df[(indicators_df["Period"] == year) & (indicators_df["indicator"].str.strip() == "At least one dose under 5-yr-old")].index
        if len(atleast_indices) > 0:
            atleast_idx = atleast_indices[0]
            atleast_df = chdn_year[chdn_year["vaccine_dose"].notna()].copy()
            if len(atleast_df) > 0:
                for quarter in quarters:
                    quarter_data = atleast_df[atleast_df["quarter"] == quarter]
                    u1_data = quarter_data[(quarter_data["age_first_num"] >= 0) & (quarter_data["age_first_num"] <= 11)]
                    u1_male = len(u1_data[u1_data["Gender"] == "Male"]["children_code(T/C Code)"].unique())
                    u1_female = len(u1_data[u1_data["Gender"] == "Female"]["children_code(T/C Code)"].unique())
                    u5_data = quarter_data[(quarter_data["age_first_num"] >= 12) & (quarter_data["age_first_num"] <= 59)]
                    u5_male = len(u5_data[u5_data["Gender"] == "Male"]["children_code(T/C Code)"].unique())
                    u5_female = len(u5_data[u5_data["Gender"] == "Female"]["children_code(T/C Code)"].unique())
                    indicators_df.loc[atleast_idx, f"{quarter} U1 Male"] = u1_male
                    indicators_df.loc[atleast_idx, f"{quarter} U1 Female"] = u1_female
                    indicators_df.loc[atleast_idx, f"{quarter} 1-5 Male "] = u5_male
                    indicators_df.loc[atleast_idx, f"{quarter} 1-5 Female"] = u5_female

        full_indices = indicators_df[(indicators_df["Period"] == year) & (indicators_df["indicator"] == "Full dose under 5-yr-old")].index
        if len(full_indices) > 0:
            full_idx = full_indices[0]
            completion_df = chdn_year[(chdn_year["completion_year"] == year) & chdn_year["completion_quarter"].notna()].copy()
            if len(completion_df) > 0:
                for quarter in quarters:
                    quarter_data = completion_df[completion_df["completion_quarter"] == quarter]
                    u1_data = quarter_data[quarter_data["completion_age"] == "U1"]
                    one5_data = quarter_data[(quarter_data["completion_age"] == "U5") | (quarter_data["completion_age"] == "1-5")]
                    u1_male = len(u1_data[u1_data["Gender"] == "Male"]["children_code(T/C Code)"].unique())
                    u1_female = len(u1_data[u1_data["Gender"] == "Female"]["children_code(T/C Code)"].unique())
                    one5_male = len(one5_data[one5_data["Gender"] == "Male"]["children_code(T/C Code)"].unique())
                    one5_female = len(one5_data[one5_data["Gender"] == "Female"]["children_code(T/C Code)"].unique())
                    indicators_df.loc[full_idx, f"{quarter} U1 Male"] = u1_male
                    indicators_df.loc[full_idx, f"{quarter} U1 Female"] = u1_female
                    indicators_df.loc[full_idx, f"{quarter} 1-5 Male "] = one5_male
                    indicators_df.loc[full_idx, f"{quarter} 1-5 Female"] = one5_female

        td_alod_indices = indicators_df[(indicators_df["Period"] == year) & (indicators_df["indicator"] == "Td ALOD")].index
        if len(td_alod_indices) > 0:
            td_alod_idx = td_alod_indices[0]
            if len(td_chdn_year) > 0:
                for quarter in quarters:
                    quarter_data = td_chdn_year[td_chdn_year["quarter"] == quarter]
                    td_alod_count = quarter_data["Pregnance_code"].nunique()
                    indicators_df.loc[td_alod_idx, f"{quarter} 1-5 Female"] = td_alod_count

        td2_indices = indicators_df[(indicators_df["Period"] == year) & (indicators_df["indicator"] == "Td Two Doses")].index
        if len(td2_indices) > 0:
            td2_idx = td2_indices[0]
            td2_df = td_chdn_year[td_chdn_year["Doses"] == "Td2"].copy()
            if len(td2_df) > 0:
                for quarter in quarters:
                    quarter_data = td2_df[td2_df["quarter"] == quarter]
                    td2_count = quarter_data["Pregnance_code"].nunique()
                    indicators_df.loc[td2_idx, f"{quarter} 1-5 Female"] = td2_count

    print(f"\n{'=' * 60}")
    print("Writing to indicators...")
    print(f"{'=' * 60}")
    wb = openpyxl.load_workbook(OUTPUT_FILE)
    if "indicators" in wb.sheetnames:
        del wb["indicators"]
        print("Removed existing 'indicators' sheet")
    ws = wb.create_sheet("indicators")
    for col_idx, col_name in enumerate(indicators_df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    for row_idx, row_data in enumerate(indicators_df.itertuples(index=False, name=None), 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    wb.save(OUTPUT_FILE)
    print(f"Successfully added 'indicators' sheet with {len(indicators_df)} rows")
    print("\nCompleted!")


def step_add_multi_year_td_alod() -> None:
    """Add Td ALOD sheet."""
    print_header("Step 9: Add Td_ALOD Sheet")
    
    td_alod_source = get_sheet_source_file("Td_ALOD")
    try:
        template_df = pd.read_excel(td_alod_source, sheet_name="Td_ALOD", engine="openpyxl")
    except Exception:
        print("  No Td_ALOD template found, skipping")
        return
    
    td_df = pd.read_excel(OUTPUT_FILE, sheet_name="Td")
    chdn_td = td_df[td_df["source"] == "CHDN"].copy()
    chdn_td["year"] = pd.to_datetime(chdn_td["reporting_month"], errors="coerce").dt.year
    
    td_alod_df = template_df.copy()
    
    years_to_process = [2024, 2025, 2026]
    for year in years_to_process:
        chdn_year = chdn_td[chdn_td["year"] == year].copy()
        if len(chdn_year) > 0:
            chdn_year["Session Point"] = chdn_year["Session Point"].astype(str).str.strip()
            td_by_clinic = chdn_year.groupby("Session Point")["Pregnance_code"].nunique().reset_index(name="count")
            td_alod_count = td_by_clinic["count"].sum()
            
            year_row_idx = td_alod_df[td_alod_df["Period"] == year].index
            if len(year_row_idx) > 0:
                idx = year_row_idx[0]
                td_alod_df.loc[idx, "Annual Achievement"] = td_alod_count
            print(f"  {year}: {td_alod_count} Td ALOD")
    
    wb = openpyxl.load_workbook(OUTPUT_FILE)
    if "Td_ALOD" in wb.sheetnames:
        del wb["Td_ALOD"]
    ws = wb.create_sheet("Td_ALOD")
    
    for col_idx, col_name in enumerate(td_alod_df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    for row_idx, row_data in enumerate(td_alod_df.itertuples(index=False, name=None), 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save(OUTPUT_FILE)
    print("OK: Td_ALOD sheet created")


def step_add_multi_year_alod_cummu() -> None:
    """Add ALOD cumulative sheet."""
    print_header("Step 10: Add ALOD_cummu Sheet")

    print("Reading data...")
    child_df = pd.read_excel(OUTPUT_FILE, sheet_name="EPI-Child-long")
    alod_cummu_source = get_sheet_source_file("ALOD_cummu")
    wb_template = openpyxl.load_workbook(alod_cummu_source)
    ws_template = wb_template["ALOD_cummu"]

    template_data = []
    for row in ws_template.iter_rows(min_row=1, max_row=ws_template.max_row, values_only=True):
        template_data.append(row)
    template_df = pd.DataFrame(template_data[1:], columns=template_data[0])

    print(f"EPI-Child-long shape: {child_df.shape}")
    print(f"Template shape: {template_df.shape}")

    chdn_child = child_df[child_df["source"] == "CHDN"].copy()
    chdn_child["year"] = pd.to_datetime(chdn_child["reporting_month"], errors="coerce").dt.year
    print(f"CHDN child records: {len(chdn_child)}")

    def calculate_alod_for_year(df, year):
        year_data = df[df["year"] == year].copy()
        if len(year_data) == 0:
            print(f"  No data for year {year}")
            return None
        print(f"  Processing year {year}: {len(year_data)} records")
        clean_data = year_data[year_data["Age at first visit"].notna() & year_data["Gender"].notna()].copy()
        print(f"    Records with Age at first visit and Gender: {len(clean_data)}")

        def get_local_age_group(age):
            if pd.isna(age):
                return None
            if age <= 11:
                return "U1"
            elif age <= 59:
                return "1-5"
            return ">5"

        clean_data["age_group"] = clean_data["Age at first visit"].apply(get_local_age_group)
        clean_data["gender_upper"] = clean_data["Gender"].str.upper()
        u1_male = clean_data[(clean_data["age_group"] == "U1") & (clean_data["gender_upper"] == "MALE")]["children_code(T/C Code)"].nunique()
        u1_female = clean_data[(clean_data["age_group"] == "U1") & (clean_data["gender_upper"] == "FEMALE")]["children_code(T/C Code)"].nunique()
        one5_male = clean_data[(clean_data["age_group"] == "1-5") & (clean_data["gender_upper"] == "MALE")]["children_code(T/C Code)"].nunique()
        one5_female = clean_data[(clean_data["age_group"] == "1-5") & (clean_data["gender_upper"] == "FEMALE")]["children_code(T/C Code)"].nunique()
        print(f"    U1 Male: {u1_male}, U1 Female: {u1_female}, 1-5 Male: {one5_male}, 1-5 Female: {one5_female}")
        return {
            "year": year,
            "u1_male": u1_male,
            "u1_female": u1_female,
            "one5_male": one5_male,
            "one5_female": one5_female,
        }

    print("\nCalculating ALOD achievements by year...")
    years_to_process = [2024, 2025, 2026]
    results = []
    for year in years_to_process:
        result = calculate_alod_for_year(chdn_child, year)
        if result:
            results.append(result)

    print(f"\nCreating ALOD_cummu sheet with {len(template_df)} rows...")
    alod_cummu_list = []
    for _, row in template_df.iterrows():
        year = int(row["Period"])
        row_copy = row.copy()
        result = calculate_alod_for_year(chdn_child, year)
        if result:
            row_copy["Annual U1 Male"] = result["u1_male"]
            row_copy["Annaul U1 Female"] = result["u1_female"]
            row_copy["Annual 1-5 Male "] = result["one5_male"]
            row_copy["Annual 1-5 Female"] = result["one5_female"]
        alod_cummu_list.append(row_copy)

    alod_cummu_df = pd.DataFrame(alod_cummu_list)
    print("\nUpdated ALOD_cummu sheet:")
    print(alod_cummu_df.to_string())

    wb = openpyxl.load_workbook(OUTPUT_FILE)
    if "ALOD_cummu" in wb.sheetnames:
        del wb["ALOD_cummu"]
        print("  Removed existing 'ALOD_cummu' sheet")
    ws = wb.create_sheet("ALOD_cummu")
    for col_idx, col_name in enumerate(alod_cummu_df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    for row_idx, row_data in enumerate(alod_cummu_df.itertuples(index=False, name=None), 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    wb.save(OUTPUT_FILE)
    print(f"Successfully added 'ALOD_cummu' sheet with {len(alod_cummu_df)} rows")

    print("\nSummary:")
    for result in results:
        year = result["year"]
        print(f"\n  Year {year}:")
        print(f"    Annual U1 Male: {result['u1_male']}")
        print(f"    Annual U1 Female: {result['u1_female']}")
        print(f"    Annual 1-5 Male: {result['one5_male']}")
        print(f"    Annual 1-5 Female: {result['one5_female']}")


def step_add_multi_year_idp() -> None:
    """Add IDP sheet."""
    print_header("Step 11: Add IDP Sheet")

    print("Reading source data and template...")
    child_df = pd.read_excel(OUTPUT_FILE, sheet_name="EPI-Child-long")
    idp_source = get_sheet_source_file("IDP")
    template_df = pd.read_excel(idp_source, sheet_name="IDP", engine="openpyxl")

    print(f"EPI-Child-long shape: {child_df.shape}")
    print(f"Template shape: {template_df.shape}")

    child_df["reporting_month"] = pd.to_datetime(child_df["reporting_month"], errors="coerce")
    child_df["period"] = child_df["reporting_month"].apply(get_period)
    child_df["gender_norm"] = child_df["Gender"].astype(str).str.strip().str.lower()
    child_df["idp_norm"] = child_df["IDP (Y/N)"].astype(str).str.strip().str.lower()
    child_df["age_at_dose_num"] = pd.to_numeric(child_df["age_at_dose"], errors="coerce")

    print("\nProcessing years from template...")
    years_to_process = template_df["Period "].unique()
    print(f"Years in template: {sorted(years_to_process)}")

    output_rows = []
    for year in sorted(years_to_process):
        print(f"\n  Processing year {year}...")
        quarters = [f"Q{i}_{year}" for i in range(1, 5)]
        filtered = child_df[
            (child_df["vaccine_dose"] == "Penta1")
            & (child_df["source"] == "CHDN")
            & (child_df["age_at_dose_num"] >= 2)
            & (child_df["age_at_dose_num"] <= 59)
            & (child_df["period"].isin(quarters))
        ].copy()
        print(f"    Filtered records (Penta1, CHDN, age 2-59, {year} quarters): {len(filtered)}")

        def count_unique(df, quarter, idp_flag, gender):
            subset = df[(df["period"] == quarter) & (df["idp_norm"] == idp_flag) & (df["gender_norm"] == gender)]
            return subset["children_code(T/C Code)"].nunique()

        results = []
        for q in quarters:
            row_counts = {
                "quarter": q,
                "idp_male": count_unique(filtered, q, "yes", "male"),
                "idp_female": count_unique(filtered, q, "yes", "female"),
                "non_idp_male": count_unique(filtered, q, "no", "male"),
                "non_idp_female": count_unique(filtered, q, "no", "female"),
            }
            print(
                f"      {q}: IDP M={row_counts['idp_male']}, IDP F={row_counts['idp_female']}, "
                f"non-IDP M={row_counts['non_idp_male']}, non-IDP F={row_counts['non_idp_female']}"
            )
            results.append(row_counts)

        template_year_rows = template_df[template_df["Period "] == year]
        if len(template_year_rows) > 0:
            base_row = template_year_rows.iloc[0].to_dict()
        else:
            base_row = {
                "Period ": year,
                "Organization ": "CHDN",
                "Project Name": "REACH_KK",
                "indicator": "Penta1 under 5-yr-old",
            }

        output_row = base_row.copy()
        output_row.setdefault("Period ", year)
        output_row.setdefault("Organization ", "CHDN")
        output_row.setdefault("Project Name", "REACH_KK")
        output_row.setdefault("indicator", base_row.get("indicator", "Penta1 under 5-yr-old"))

        for res in results:
            q_short = res["quarter"].split("_")[0]
            output_row[f"{q_short} IDP Male "] = res["idp_male"] if res["idp_male"] > 0 else None
            output_row[f"{q_short} IDP Female"] = res["idp_female"] if res["idp_female"] > 0 else None
            output_row[f"{q_short} non-IDP Male "] = res["non_idp_male"] if res["non_idp_male"] > 0 else None
            output_row[f"{q_short} non-IDP Female"] = res["non_idp_female"] if res["non_idp_female"] > 0 else None

        output_rows.append(output_row)

    out_df = pd.DataFrame(output_rows)
    for col in template_df.columns:
        if col not in out_df.columns:
            out_df[col] = None
    out_df = out_df[template_df.columns]

    print("\nFinal IDP rows:")
    print(out_df.to_string(index=False))

    wb = openpyxl.load_workbook(OUTPUT_FILE)
    if "IDP" in wb.sheetnames:
        del wb["IDP"]
        print("Removed existing 'IDP' sheet")
    ws = wb.create_sheet("IDP")
    for col_idx, col_name in enumerate(out_df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    for row_idx, row_data in enumerate(out_df.itertuples(index=False, name=None), 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    wb.save(OUTPUT_FILE)
    print(f"\nAdded 'IDP' sheet to {OUTPUT_FILE.name} with {len(out_df)} rows")


def step_add_multi_year_td2_indicator() -> None:
    """Add Td2_indicator sheet."""
    print_header("Step 12: Add Td2_indicator Sheet")

    print("\n1. Reading Td2_indicator template...")
    td2_source = get_sheet_source_file("Td2_indicator")
    template_df = pd.read_excel(td2_source, sheet_name="Td2_indicator", engine="openpyxl")
    years_to_process = sorted(template_df["Period"].unique())
    print(f"   Years in template: {years_to_process}")

    print("\n2. Reading Td sheet from CHDN dataset_long.xlsx...")
    td_df = pd.read_excel(OUTPUT_FILE, sheet_name="Td", engine="openpyxl")
    print(f"   Td sheet shape: {td_df.shape}")

    print("\n3. Filtering for Doses='Td2' and source='CHDN'...")
    td2_df = td_df[(td_df["Doses"] == "Td2") & (td_df["source"] == "CHDN")].copy()
    print(f"   Filtered data shape: {td2_df.shape}")

    def get_quarter(date):
        if pd.isna(date):
            return None
        d = pd.to_datetime(date)
        y = d.year
        q1_start = pd.Timestamp(year=y - 1, month=12, day=21)
        q1_end = pd.Timestamp(year=y, month=3, day=20)
        if q1_start <= d <= q1_end:
            return "Q1"
        q2_start = pd.Timestamp(year=y, month=3, day=21)
        q2_end = pd.Timestamp(year=y, month=6, day=20)
        if q2_start <= d <= q2_end:
            return "Q2"
        q3_start = pd.Timestamp(year=y, month=6, day=21)
        q3_end = pd.Timestamp(year=y, month=9, day=20)
        if q3_start <= d <= q3_end:
            return "Q3"
        q4_start = pd.Timestamp(year=y, month=9, day=21)
        q4_end = pd.Timestamp(year=y, month=12, day=20)
        if q4_start <= d <= q4_end:
            return "Q4"
        return "Q1"

    td2_df["reporting_month"] = pd.to_datetime(td2_df["reporting_month"])
    td2_df["year"] = td2_df["reporting_month"].dt.year
    td2_df["quarter"] = td2_df["reporting_month"].apply(get_quarter)

    print("\n4. Calculating quarterly achievements for each year...")
    yearly_results = []
    for year in years_to_process:
        print(f"\n   Year {year}:")
        td2_year = td2_df[(td2_df["year"] == year) & td2_df["quarter"].notna()].copy()
        print(f"     Records for {year}: {len(td2_year)}")
        if td2_year.empty:
            print(f"     No Td2 data found for {year}")
            quarterly_counts = {"Q1": 0, "Q2": 0, "Q3": 0, "Q4": 0}
        else:
            quarterly_counts = {}
            for quarter in ["Q1", "Q2", "Q3", "Q4"]:
                q_df = td2_year[td2_year["quarter"] == quarter]
                count = q_df["Pregnance_code"].nunique()
                quarterly_counts[quarter] = count
                print(f"     {quarter}: {count} unique Pregnance_code with Td2")
        yearly_results.append({
            "year": year,
            "q1": quarterly_counts["Q1"],
            "q2": quarterly_counts["Q2"],
            "q3": quarterly_counts["Q3"],
            "q4": quarterly_counts["Q4"],
        })

    print("\n5. Creating Td2_indicator sheet...")
    indicator_rows = []
    for result in yearly_results:
        year = result["year"]
        template_year = template_df[template_df["Period"] == year]
        if len(template_year) > 0:
            row = template_year.iloc[0].to_dict()
        else:
            row = {
                "Period": year,
                "Organization ": "CHDN",
                "Project Name": "REACH-KK",
                "Indicators ": "Td Two Doses",
            }
        row["Q1 Achievement"] = result["q1"]
        row["Q2 Achievement"] = result["q2"]
        row["Q3 Achievement"] = result["q3"]
        row["Q4 Achievement"] = result["q4"]
        indicator_rows.append(row)

    indicator_df = pd.DataFrame(indicator_rows)
    for col in template_df.columns:
        if col not in indicator_df.columns:
            indicator_df[col] = None
    indicator_df = indicator_df[template_df.columns]

    print("   Td2_indicator data:")
    print(indicator_df.to_string(index=False))

    print("\n6. Adding Td2_indicator sheet to CHDN dataset_long.xlsx...")
    wb = openpyxl.load_workbook(OUTPUT_FILE)
    if "Td2_indicator" in wb.sheetnames:
        print("   Removing existing Td2_indicator sheet...")
        del wb["Td2_indicator"]
    ws = wb.create_sheet("Td2_indicator")
    for col_idx, col_name in enumerate(indicator_df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    for row_idx, row_data in enumerate(indicator_df.itertuples(index=False, name=None), 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    wb.save(OUTPUT_FILE)
    print("   OK: Td2_indicator sheet added successfully!")


def main() -> int:
    """Run all transformation steps."""
    parser = argparse.ArgumentParser(
        description="Run full CHDN EPI transformation pipeline (single consolidated script)."
    )
    args = parser.parse_args()

    # Resolve data file (use fallback if primary is locked)
    global DATA_FILE
    DATA_FILE = get_data_file()
    
    try:
        initialize_output_file()
        step_transform_epi_long()
        step_transform_pregnancy_long()
        step_verify_output()
        step_create_summary_sheet()
        step_add_alod_columns()
        step_add_vaccine_columns()
        step_add_cumulative_sheet()
        step_add_cumulative()
        step_add_multi_year_indicators()
        step_add_multi_year_td_alod()
        step_add_multi_year_alod_cummu()
        step_add_multi_year_idp()
        step_add_multi_year_td2_indicator()

        print_header("SUCCESS: All transformations complete!")
        print("Output file: CHDN dataset_long.xlsx")
        print("Sheets generated:")
        print(" - EPI-Child-long")
        print(" - Td")
        print(" - Summary")
        print(" - yearly_cumulative")
        print(" - cumulative")
        print(" - indicators")
        print(" - Td_ALOD")
        print(" - ALOD_cummu")
        print(" - IDP")
        print(" - Td2_indicator")

        return 0

    except Exception as e:
        print(f"\nERROR: {str(e)}")
        import traceback

        traceback.print_exc()
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
