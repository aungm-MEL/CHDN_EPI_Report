import argparse
import sys
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


DEFAULT_INPUT = Path(__file__).resolve().with_name("EPI Database_CHDN.xlsx")
DEFAULT_OUTPUT = Path(__file__).resolve().with_name("CHDN_EPI_clean.xlsx")
DEFAULT_SOURCE_SHEET = "EPI-Child"
DEFAULT_TARGET_SHEET = "EPI-Child"
DEFAULT_PREGNANCY_SOURCE_SHEET = "EPI-Pregnancy"
DEFAULT_PREGNANCY_TARGET_SHEET = "EPI-Pregnancy"

COMPLETE_COLUMNS = [
    "CompleteInQ4 2024",
    "CompleteInQ1 2025",
    "CompleteInQ2 2025",
    "CompleteInQ3 2025",
    "CompleteInQ4 2025",
    "CompleteInQ1 2026",
    "CompleteInQ2 2026",
    "CompleteInQ3 2026",
    "CompleteInQ4 2026",
]

QUARTER_WINDOWS = {
    "CompleteInQ4 2024": ("Q4", date(2024, 9, 21), date(2024, 12, 20)),
    "CompleteInQ1 2025": ("Q1", date(2024, 12, 21), date(2025, 3, 20)),
    "CompleteInQ2 2025": ("Q2", date(2025, 3, 21), date(2025, 6, 20)),
    "CompleteInQ3 2025": ("Q3", date(2025, 6, 21), date(2025, 9, 20)),
    "CompleteInQ4 2025": ("Q4", date(2025, 9, 21), date(2025, 12, 20)),
    "CompleteInQ1 2026": ("Q1", date(2025, 12, 21), date(2026, 3, 20)),
    "CompleteInQ2 2026": ("Q2", date(2026, 3, 21), date(2026, 6, 20)),
    "CompleteInQ3 2026": ("Q3", date(2026, 6, 21), date(2026, 9, 20)),
    "CompleteInQ4 2026": ("Q4", date(2026, 9, 21), date(2026, 12, 20)),
}


def _normalize_col_name(name: str) -> str:
    return "".join(ch for ch in str(name).strip().lower() if ch.isalnum())


def _find_children_code_column(columns) -> int | None:
    aliases = {
        "childrencodetccode",
        "childrencode",
        "tccode",
        "tcode",
    }
    for idx, col in enumerate(columns, start=1):
        if _normalize_col_name(col) in aliases:
            return idx
    return None


def _find_pregnance_code_column(columns) -> int | None:
    aliases = {
        "pregnancecode",
        "pregnancycode",
        "pregnance_code",
        "pregnancy_code",
        "pwcode",
        "pw_code",
    }
    norm_aliases = {a.replace("_", "") for a in aliases}
    for idx, col in enumerate(columns, start=1):
        if _normalize_col_name(col) in norm_aliases:
            return idx
    return None


def _is_blank(value) -> bool:
    return value is None or str(value).strip() == ""


def _is_formula_value(value) -> bool:
    return isinstance(value, str) and value.strip().startswith("=")


def _is_yes(value) -> bool:
    return str(value).strip().upper() == "YES"


def _to_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        parsed = pd.to_datetime(value, errors="coerce")
        if pd.isna(parsed):
            return None
        return parsed.date()
    if isinstance(value, (int, float)):
        try:
            parsed = pd.to_datetime(value, unit="D", origin="1899-12-30", errors="coerce")
            if pd.isna(parsed):
                return None
            return parsed.date()
        except Exception:
            return None
    return None


def _row_value(row_values, col_letter: str):
    return row_values[column_index_from_string(col_letter) - 1]


def _presence_ok(row_values, dose_col: str, other_col: str) -> bool:
    return (not _is_blank(_row_value(row_values, dose_col))) or _is_yes(_row_value(row_values, other_col))


def _date_source_ok(row_values, date_col: str, source_col: str, start: date, end: date) -> bool:
    row_date = _to_date(_row_value(row_values, date_col))
    if row_date is None or not (start <= row_date <= end):
        return False
    return str(_row_value(row_values, source_col)).strip().upper() == "CHDN"


def _max_date_ok(row_values, cols: list[str], end: date) -> bool:
    dates = [_to_date(_row_value(row_values, col)) for col in cols]
    dates = [item for item in dates if item is not None]
    if not dates:
        return True
    return max(dates) <= end


def _completion_value(row_values, column_name: str) -> str:
    quarter_label, start_date, end_date = QUARTER_WINDOWS[column_name]
    year_label = column_name.rsplit(" ", 1)[-1]

    u1_presence_pairs = [
        ("L", "M"),
        ("Q", "R"),
        ("V", "W"),
        ("AA", "AB"),
        ("AF", "AG"),
        ("AK", "AL"),
        ("AP", "AQ"),
        ("AU", "AV"),
    ]
    u5_presence_pairs = u1_presence_pairs[1:]

    u1_date_source_pairs = [
        ("N", "P"),
        ("S", "U"),
        ("X", "Z"),
        ("AC", "AE"),
        ("AH", "AJ"),
        ("AM", "AO"),
        ("AR", "AT"),
        ("AW", "AY"),
        ("BB", "BD"),
    ]
    u5_date_source_pairs = u1_date_source_pairs[1:]

    age_value = _row_value(row_values, "CE")
    try:
        age_num = float(age_value)
    except (TypeError, ValueError):
        age_num = None

    if age_num is not None and age_num <= 11:
        if all(_presence_ok(row_values, dose_col, other_col) for dose_col, other_col in u1_presence_pairs):
            if any(_date_source_ok(row_values, date_col, source_col, start_date, end_date) for date_col, source_col in u1_date_source_pairs):
                if _max_date_ok(row_values, ["AC", "AR", "AW"], end_date):
                    return f"U1 complete in {quarter_label}_{year_label}"

    if age_num is not None and 11 < age_num <= 59:
        if all(_presence_ok(row_values, dose_col, other_col) for dose_col, other_col in u5_presence_pairs):
            if any(_date_source_ok(row_values, date_col, source_col, start_date, end_date) for date_col, source_col in u5_date_source_pairs):
                if _max_date_ok(row_values, ["AC", "AR", "AW"], end_date):
                    return f"1-5 complete in {quarter_label}_{year_label}"

    return ""


def _get_kept_rows(values_sheet, formula_sheet, data_start_row: int, max_col: int) -> tuple[list[int], int]:
    kept_rows: list[int] = []
    removed_count = 0
    max_row = max(values_sheet.max_row, formula_sheet.max_row)

    for row_num in range(data_start_row, max_row + 1):
        has_formula = False
        has_non_formula_data = False

        for col_idx in range(1, max_col + 1):
            formula_cell_val = formula_sheet.cell(row=row_num, column=col_idx).value
            value_cell_val = values_sheet.cell(row=row_num, column=col_idx).value

            if _is_formula_value(formula_cell_val):
                has_formula = True
            elif not _is_blank(value_cell_val):
                has_non_formula_data = True

        if not has_non_formula_data:
            removed_count += 1
        else:
            kept_rows.append(row_num)

    return kept_rows, removed_count


def _warn_code_quality(sheet, code_col_idx: int, code_label: str, row_numbers: list[int]) -> None:
    code_values = [sheet.cell(row=row_num, column=code_col_idx).value for row_num in row_numbers]
    normalized_codes = [str(value).strip() for value in code_values if not _is_blank(value)]
    missing_count = sum(1 for value in code_values if _is_blank(value))
    duplicate_codes = sorted({value for value in normalized_codes if normalized_codes.count(value) > 1})

    if missing_count > 0:
        print(f"WARNING: {code_label} missing. Missing count = {missing_count}")
    else:
        print(f"No missing {code_label}")
    if duplicate_codes:
        print(f"WARNING: duplication of {code_label} ({','.join(duplicate_codes)})")


def _write_sheet_values(src_ws, dst_ws) -> None:
    for row_num in range(1, src_ws.max_row + 1):
        for col_idx in range(1, src_ws.max_column + 1):
            dst_ws.cell(row=row_num, column=col_idx, value=src_ws.cell(row=row_num, column=col_idx).value)


def _write_rows_sheet(out_wb, sheet_name: str, columns: list[str], rows: list[dict]) -> None:
    if sheet_name in out_wb.sheetnames:
        del out_wb[sheet_name]
    ws = out_wb.create_sheet(title=sheet_name)

    for col_idx, col_name in enumerate(columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(col_name))


def _add_default_template_sheets(out_wb) -> list[str]:
    years = [2024, 2025, 2026]
    created: list[str] = []

    indicator_names = [
        "Penta3 under 1-yr-old",
        "Penta3 under 5-yr-old",
        "MMR1 under 1-yr-old",
        "MMR1 under 5-yr-old",
        "MMR2 under 5-yr-old",
        "Penta1 under 5-yr-old",
        "At least one dose under 5-yr-old",
        "Full dose under 5-yr-old",
        "Td ALOD",
        "Td Two Doses",
    ]
    quarter_cols = [
        "Q1 U1 Male",
        "Q1 U1 Female",
        "Q1 1-5 Male ",
        "Q1 1-5 Female",
        "Q2 U1 Male",
        "Q2 U1 Female",
        "Q2 1-5 Male ",
        "Q2 1-5 Female",
        "Q3 U1 Male",
        "Q3 U1 Female",
        "Q3 1-5 Male ",
        "Q3 1-5 Female",
        "Q4 U1 Male",
        "Q4 U1 Female",
        "Q4 1-5 Male ",
        "Q4 1-5 Female",
    ]
    indicator_columns = ["Period", "Organization", "Project Name", "indicator"] + quarter_cols
    indicator_rows = []
    for year in years:
        for name in indicator_names:
            row = {
                "Period": year,
                "Organization": "CHDN",
                "Project Name": "REACH-KK",
                "indicator": name,
            }
            for col in quarter_cols:
                row[col] = None
            indicator_rows.append(row)
    _write_rows_sheet(out_wb, "indicators", indicator_columns, indicator_rows)
    created.append("indicators")

    td_alod_columns = ["Period", "Organization ", "Project Name", "Indicators ", "Annual Achievement"]
    td_alod_rows = [
        {
            "Period": year,
            "Organization ": "CHDN",
            "Project Name": "REACH-KK",
            "Indicators ": "Td ALOD",
            "Annual Achievement": None,
        }
        for year in years
    ]
    _write_rows_sheet(out_wb, "Td_ALOD", td_alod_columns, td_alod_rows)
    created.append("Td_ALOD")

    alod_cummu_columns = [
        "Period",
        "Organization ",
        "Project Name",
        "Indicator",
        "Annual U1 Male",
        "Annaul U1 Female",
        "Annual 1-5 Male ",
        "Annual 1-5 Female",
    ]
    alod_cummu_rows = [
        {
            "Period": year,
            "Organization ": "CHDN",
            "Project Name": "REACH-KK",
            "Indicator": "ALOD cumulative",
            "Annual U1 Male": None,
            "Annaul U1 Female": None,
            "Annual 1-5 Male ": None,
            "Annual 1-5 Female": None,
        }
        for year in years
    ]
    _write_rows_sheet(out_wb, "ALOD_cummu", alod_cummu_columns, alod_cummu_rows)
    created.append("ALOD_cummu")

    idp_columns = [
        "Period ",
        "Organization ",
        "Project Name",
        "indicator",
        "Q1 IDP Male ",
        "Q1 IDP Female",
        "Q1 non-IDP Male ",
        "Q1 non-IDP Female",
        "Q2 IDP Male ",
        "Q2 IDP Female",
        "Q2 non-IDP Male ",
        "Q2 non-IDP Female",
        "Q3 IDP Male ",
        "Q3 IDP Female",
        "Q3 non-IDP Male ",
        "Q3 non-IDP Female",
        "Q4 IDP Male ",
        "Q4 IDP Female",
        "Q4 non-IDP Male ",
        "Q4 non-IDP Female",
    ]
    idp_rows = [
        {
            "Period ": year,
            "Organization ": "CHDN",
            "Project Name": "REACH-KK",
            "indicator": "Penta1 under 5-yr-old",
        }
        for year in years
    ]
    _write_rows_sheet(out_wb, "IDP", idp_columns, idp_rows)
    created.append("IDP")

    td2_columns = [
        "Period",
        "Organization ",
        "Project Name",
        "Indicators ",
        "Q1 Achievement",
        "Q2 Achievement",
        "Q3 Achievement",
        "Q4 Achievement",
    ]
    td2_rows = [
        {
            "Period": year,
            "Organization ": "CHDN",
            "Project Name": "REACH-KK",
            "Indicators ": "Td Two Doses",
            "Q1 Achievement": None,
            "Q2 Achievement": None,
            "Q3 Achievement": None,
            "Q4 Achievement": None,
        }
        for year in years
    ]
    _write_rows_sheet(out_wb, "Td2_indicator", td2_columns, td2_rows)
    created.append("Td2_indicator")

    return created


def build_clean_sheet(
    input_file: Path,
    output_file: Path,
    source_sheet_name: str,
    target_sheet_name: str,
    pregnancy_source_sheet_name: str,
    pregnancy_target_sheet_name: str,
) -> int:
    if not input_file.exists():
        print(f"ERROR: Input file not found: {input_file}")
        return 2

    try:
        wb = load_workbook(input_file, data_only=True)
        wb_formula = load_workbook(input_file, data_only=False)
    except Exception as exc:
        print(f"ERROR: Failed to open input workbook: {exc}")
        return 2

    if source_sheet_name not in wb.sheetnames:
        print(f"ERROR: Source sheet not found: {source_sheet_name}")
        return 2

    child_ws = wb[source_sheet_name]
    child_formula_ws = wb_formula[source_sheet_name]
    max_col = child_ws.max_column

    source_headers = [child_ws.cell(row=1, column=idx).value for idx in range(1, max_col + 1)]
    code_col_idx = _find_children_code_column(source_headers)
    if code_col_idx is None:
        print("ERROR: Could not find children_code(T/C Code) column in the source sheet.")
        print(f"Available columns: {', '.join(map(str, source_headers))}")
        return 2

    child_rows_to_keep, child_removed_count = _get_kept_rows(
        values_sheet=child_ws,
        formula_sheet=child_formula_ws,
        data_start_row=2,
        max_col=max_col,
    )
    if child_removed_count > 0:
        print(f"INFO: Removed {child_removed_count} formula-only rows from {source_sheet_name}")

    _warn_code_quality(child_ws, code_col_idx, "children_code", child_rows_to_keep)

    out_wb = Workbook()
    child_out_ws = out_wb.active
    child_out_ws.title = target_sheet_name

    for col_idx, header in enumerate(source_headers, start=1):
        child_out_ws.cell(row=1, column=col_idx, value=header)

    for extra_idx, col_name in enumerate(COMPLETE_COLUMNS, start=max_col + 1):
        child_out_ws.cell(row=1, column=extra_idx, value=col_name)

    out_child_row = 2
    for row_num in child_rows_to_keep:
        row_values = [child_ws.cell(row=row_num, column=col_idx).value for col_idx in range(1, max_col + 1)]
        for col_idx, value in enumerate(row_values, start=1):
            child_out_ws.cell(row=out_child_row, column=col_idx, value=value)

        for extra_idx, col_name in enumerate(COMPLETE_COLUMNS, start=max_col + 1):
            child_out_ws.cell(row=out_child_row, column=extra_idx, value=_completion_value(row_values, col_name))

        out_child_row += 1

    if pregnancy_source_sheet_name in wb.sheetnames:
        preg_ws = wb[pregnancy_source_sheet_name]
        preg_formula_ws = wb_formula[pregnancy_source_sheet_name]
        preg_max_col = preg_ws.max_column
        preg_header_row = 2
        preg_headers = [preg_ws.cell(row=preg_header_row, column=idx).value for idx in range(1, preg_max_col + 1)]

        preg_rows_to_keep, preg_removed_count = _get_kept_rows(
            values_sheet=preg_ws,
            formula_sheet=preg_formula_ws,
            data_start_row=3,
            max_col=preg_max_col,
        )
        if preg_removed_count > 0:
            print(f"INFO: Removed {preg_removed_count} formula-only rows from {pregnancy_source_sheet_name}")

        preg_code_col_idx = _find_pregnance_code_column(preg_headers)
        if preg_code_col_idx is None:
            print("WARNING: Could not find Pregnance_code column in EPI-Pregnancy sheet.")
            print(f"Available columns: {', '.join(map(str, preg_headers))}")
        else:
            _warn_code_quality(preg_ws, preg_code_col_idx, "pw_code", preg_rows_to_keep)

        preg_out_ws = out_wb.create_sheet(title=pregnancy_target_sheet_name)
        for col_idx in range(1, preg_max_col + 1):
            preg_out_ws.cell(row=1, column=col_idx, value=preg_ws.cell(row=preg_header_row, column=col_idx).value)

        out_preg_row = 2
        for row_num in preg_rows_to_keep:
            for col_idx in range(1, preg_max_col + 1):
                preg_out_ws.cell(row=out_preg_row, column=col_idx, value=preg_ws.cell(row=row_num, column=col_idx).value)
            out_preg_row += 1
    else:
        print(f"WARNING: Pregnancy source sheet not found: {pregnancy_source_sheet_name}")

    created_templates = _add_default_template_sheets(out_wb)

    try:
        out_wb.save(output_file)
    except Exception as exc:
        print(f"ERROR: Failed to save workbook: {exc}")
        return 2

    print(f"Done: created {output_file.name}")
    print(f"Sheets: {target_sheet_name}, {pregnancy_target_sheet_name}")
    if created_templates:
        print("Created template sheets: " + ", ".join(created_templates))
    print("Added columns: " + ", ".join(COMPLETE_COLUMNS))

    return 0


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Create CHDN_EPI_clean.xlsx with CompleteIn quarter columns and copy all other input sheets"
    )
    parser.add_argument("--input", default=str(DEFAULT_INPUT), help="Input Excel workbook path")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT), help="Output Excel workbook path")
    parser.add_argument("--source-sheet", default=DEFAULT_SOURCE_SHEET, help="Source sheet name")
    parser.add_argument("--target-sheet", default=DEFAULT_TARGET_SHEET, help="New target sheet name")
    parser.add_argument("--preg-source-sheet", default=DEFAULT_PREGNANCY_SOURCE_SHEET, help="Pregnancy source sheet name")
    parser.add_argument("--preg-target-sheet", default=DEFAULT_PREGNANCY_TARGET_SHEET, help="Pregnancy target sheet name")
    args = parser.parse_args()

    return build_clean_sheet(
        input_file=Path(args.input),
        output_file=Path(args.output),
        source_sheet_name=args.source_sheet,
        target_sheet_name=args.target_sheet,
        pregnancy_source_sheet_name=args.preg_source_sheet,
        pregnancy_target_sheet_name=args.preg_target_sheet,
    )


if __name__ == "__main__":
    sys.exit(main())


